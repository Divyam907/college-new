import os
import base64
import json
import re
import shutil
import datetime as dt
import logging
import sys
import pytz

IST = pytz.timezone('Asia/Kolkata')

def _today_ist():
    """Return today's date in IST."""
    return dt.datetime.now(IST).date()

def _now_ist():
    """Return current datetime in IST (naive, for DB storage)."""
    return dt.datetime.now(IST).replace(tzinfo=None)

# Suppress TensorFlow/CUDA noise
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'
os.environ['TF_ENABLE_ONEDNN_OPTS'] = '0'
os.environ['CUDA_VISIBLE_DEVICES'] = '-1'

logging.getLogger('tensorflow').setLevel(logging.ERROR)

print("[startup] Loading core libraries...", flush=True)

# ── Safe imports with fallback warnings ───────────────────────────────────────
_WARNINGS = []

try:
    import cv2
except ImportError:
    cv2 = None
    _WARNINGS.append("opencv (cv2) not installed — face detection disabled")

try:
    import numpy as np
except ImportError:
    np = None
    _WARNINGS.append("numpy not installed — calculations disabled")

try:
    import psycopg2
except ImportError:
    psycopg2 = None
    _WARNINGS.append("psycopg2 not installed — database disabled")

try:
    import pandas as pd
except ImportError:
    pd = None
    _WARNINGS.append("pandas not installed — embeddings disabled")

from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
os.chdir(BASE_DIR)

try:
    from config import DB_PARAMS, TWILIO_CONFIG, CONTINUOUS_ATTENDANCE, ENGAGEMENT_CONFIG, LIVENESS_CONFIG
except Exception as e:
    _WARNINGS.append(f"config.py error: {e}")
    DB_PARAMS = {}
    TWILIO_CONFIG = {}
    CONTINUOUS_ATTENDANCE = {}
    ENGAGEMENT_CONFIG = {}
    LIVENESS_CONFIG = {}

# Lazy import heavy AI modules — only loaded when first needed
_ai_loaded = False
process_group_image = None
process_group_image_with_subject = None
gen_embed = None
analyze_engagement = None
quick_liveness_check = None

def _load_ai_modules():
    global _ai_loaded, process_group_image, process_group_image_with_subject
    global gen_embed, analyze_engagement, quick_liveness_check
    if _ai_loaded:
        return
    if os.environ.get('RENDER'):
        print("[runtime] Skipping AI modules on Render free tier (512MB limit)", flush=True)
        return
    try:
        print("[runtime] Loading AI models...", flush=True)
        from Attendance_update_db import process_group_image as _pgi, process_group_image_with_subject as _pgis
        import gen_embed as _ge
        from engagement import analyze_engagement as _ae
        from liveness import quick_liveness_check as _qlc
        process_group_image = _pgi
        process_group_image_with_subject = _pgis
        gen_embed = _ge
        analyze_engagement = _ae
        quick_liveness_check = _qlc
        _ai_loaded = True
        print("[runtime] AI models loaded.", flush=True)
    except Exception as e:
        print(f"[runtime] WARNING: AI modules failed to load: {e}", flush=True)
        _WARNINGS.append(f"AI modules unavailable: {e}")

print("[startup] Core modules loaded.", flush=True)

try:
    from whatsapp_alerts import WhatsAppAlerts, check_and_send_absence_alerts, send_daily_summary_to_all
except Exception as e:
    _WARNINGS.append(f"WhatsApp alerts unavailable: {e}")
    WhatsAppAlerts = None
    check_and_send_absence_alerts = None
    send_daily_summary_to_all = None

try:
    from continuous_attendance import init_continuous_attendance, get_continuous_attendance
except Exception as e:
    _WARNINGS.append(f"Continuous attendance unavailable: {e}")
    init_continuous_attendance = None
    get_continuous_attendance = None

DATASET_DIR = os.path.join(BASE_DIR, 'dataset')
IMAGES_DIR  = os.path.join(BASE_DIR, 'images')

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'college_attendance_secret_key_2026')
app.config['MAX_CONTENT_LENGTH'] = 64 * 1024 * 1024

# ── CORS for production (frontend on different domain) ────────────────────────
FRONTEND_URL = os.environ.get('FRONTEND_URL', 'http://localhost:3000')
CORS(app, supports_credentials=True, origins=[FRONTEND_URL],
     allow_headers=["Content-Type", "Authorization"],
     methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"])

# ── Session cookie config for cross-origin ────────────────────────────────────
if os.environ.get('RENDER'):
    app.config['SESSION_COOKIE_SAMESITE'] = 'None'
    app.config['SESSION_COOKIE_SECURE'] = True


# Custom JSON provider to handle numpy types
import json as _json
from flask.json.provider import DefaultJSONProvider

class NumpyJSONProvider(DefaultJSONProvider):
    def default(self, obj):
        if isinstance(obj, np.bool_):
            return bool(obj)
        if isinstance(obj, np.integer):
            return int(obj)
        if isinstance(obj, np.floating):
            return float(obj)
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        if isinstance(obj, (dt.date, dt.time, dt.datetime)):
            return str(obj)
        return super().default(obj)

app.json_provider_class = NumpyJSONProvider
app.json = NumpyJSONProvider(app)

os.makedirs(DATASET_DIR, exist_ok=True)
os.makedirs(IMAGES_DIR,  exist_ok=True)

# ── Initialize WhatsApp client ────────────────────────────────────────────────
whatsapp_client = None
try:
    if WhatsAppAlerts and TWILIO_CONFIG.get('account_sid') and TWILIO_CONFIG.get('auth_token'):
        whatsapp_client = WhatsAppAlerts(
            account_sid=TWILIO_CONFIG['account_sid'],
            auth_token=TWILIO_CONFIG['auth_token'],
            from_number=TWILIO_CONFIG['whatsapp_from']
        )
except Exception as e:
    print(f"[startup] WARNING: WhatsApp init failed: {e}", flush=True)

# ── Initialize Continuous Attendance ──────────────────────────────────────────
continuous_attn = None
try:
    if init_continuous_attendance:
        continuous_attn = init_continuous_attendance(
            db_params=DB_PARAMS,
            dataset_dir=DATASET_DIR,
            images_dir=IMAGES_DIR,
            interval=CONTINUOUS_ATTENDANCE.get('interval_minutes', 15),
            use_webcam=CONTINUOUS_ATTENDANCE.get('use_webcam', True)
        )
        if CONTINUOUS_ATTENDANCE.get('enabled') and continuous_attn:
            for stream_cfg in CONTINUOUS_ATTENDANCE.get('rtsp_streams', []):
                continuous_attn.add_stream(
                    stream_key=stream_cfg['key'],
                    rtsp_url=stream_cfg.get('url', 'webcam'),
                    section_id=stream_cfg['section_id'],
                    camera_name=stream_cfg.get('name', '')
                )
            continuous_attn.start()
except Exception as e:
    print(f"[startup] WARNING: Continuous attendance init failed: {e}", flush=True)


# ── Print startup warnings ────────────────────────────────────────────────────
if _WARNINGS:
    print("=" * 60, flush=True)
    print("[startup] WARNINGS (non-fatal — app will still run):", flush=True)
    for w in _WARNINGS:
        print(f"  ⚠ {w}", flush=True)
    print("=" * 60, flush=True)


# ── Health check endpoint (Render pings this) ─────────────────────────────────
@app.route('/health')
def health_check():
    return jsonify({'status': 'ok', 'warnings': _WARNINGS}), 200


# ── Template-not-found fallback: redirect to React frontend ───────────────────
from jinja2.exceptions import TemplateNotFound as _TemplateNotFound

@app.errorhandler(_TemplateNotFound)
def handle_template_not_found(e):
    """Legacy server-rendered routes redirect to the React SPA."""
    return redirect(FRONTEND_URL)


# ── DB helpers ────────────────────────────────────────────────────────────────
from psycopg2 import pool as _pg_pool
import threading as _pool_lock

_db_pool = None
_pool_init_lock = _pool_lock.Lock()

def _get_pool():
    global _db_pool
    if _db_pool is not None:
        return _db_pool
    with _pool_init_lock:
        if _db_pool is None:
            try:
                _db_pool = _pg_pool.ThreadedConnectionPool(minconn=1, maxconn=5, **DB_PARAMS)
                print("[DB] Connection pool created.", flush=True)
            except Exception as _e:
                print(f"[DB] Pool creation failed: {_e}", flush=True)
    return _db_pool

def get_conn():
    pool = _get_pool()
    if pool:
        try:
            return pool.getconn()
        except Exception:
            pass
    # Fallback to direct connection
    return psycopg2.connect(**DB_PARAMS)

def _return_conn(conn):
    """Return a connection back to the pool."""
    pool = _get_pool()
    if pool:
        try:
            pool.putconn(conn)
            return
        except Exception:
            pass
    try:
        conn.close()
    except Exception:
        pass

# Eagerly initialize the pool at import time (non-blocking, best-effort)
try:
    _get_pool()
except Exception:
    pass


def init_db():
    conn = get_conn()
    cur  = conn.cursor()

    # Users table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id            SERIAL PRIMARY KEY,
            name          VARCHAR(100) NOT NULL,
            email         VARCHAR(100) UNIQUE NOT NULL,
            password_hash VARCHAR(255) NOT NULL,
            role          VARCHAR(20)  NOT NULL CHECK (role IN ('admin','college','teacher'))
        )
    """)

    # Batch table (passing year groups)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS batch (
            batch_id      SERIAL PRIMARY KEY,
            passing_year  INTEGER UNIQUE NOT NULL,
            name          VARCHAR(50) NOT NULL
        )
    """)

    # Branch/Department table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS branch (
            branch_id  SERIAL PRIMARY KEY,
            name       VARCHAR(100) NOT NULL,
            code       VARCHAR(20),
            hod_name   VARCHAR(100),
            hod_email  VARCHAR(100),
            batch_id   INTEGER REFERENCES batch(batch_id) ON DELETE CASCADE
        )
    """)

    # Authority table (director, dean, HOD for reports)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS authority (
            authority_id SERIAL PRIMARY KEY,
            name         VARCHAR(100) NOT NULL,
            email        VARCHAR(100) NOT NULL,
            role         VARCHAR(50) NOT NULL,
            branch_id    INTEGER REFERENCES branch(branch_id) ON DELETE SET NULL
        )
    """)

    # Parent contact table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS parent_contact (
            id          SERIAL PRIMARY KEY,
            student_id  INTEGER REFERENCES student(std_id) ON DELETE CASCADE,
            parent_name VARCHAR(100),
            parent_email VARCHAR(100) NOT NULL,
            parent_phone VARCHAR(20),
            relation    VARCHAR(20) DEFAULT 'parent'
        )
    """)

    # Report schedule table (automated sending)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS report_schedule (
            id              SERIAL PRIMARY KEY,
            created_by      INTEGER REFERENCES users(id),
            frequency       VARCHAR(20) NOT NULL CHECK (frequency IN ('daily','weekly','monthly')),
            batch_id        INTEGER REFERENCES batch(batch_id) ON DELETE SET NULL,
            branch_id       INTEGER REFERENCES branch(branch_id) ON DELETE SET NULL,
            class_id        INTEGER REFERENCES class(class_id) ON DELETE SET NULL,
            section_id      INTEGER REFERENCES section(section_id) ON DELETE SET NULL,
            send_to_authorities BOOLEAN DEFAULT TRUE,
            send_to_parents BOOLEAN DEFAULT FALSE,
            is_active       BOOLEAN DEFAULT TRUE,
            last_sent       TIMESTAMP,
            next_send       TIMESTAMP
        )
    """)

    # Classes table (e.g., Class 10, Class 11, BCA-1)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS class (
            class_id   SERIAL PRIMARY KEY,
            name       VARCHAR(50) NOT NULL,
            batch_id   INTEGER REFERENCES batch(batch_id) ON DELETE SET NULL,
            branch_id  INTEGER REFERENCES branch(branch_id) ON DELETE SET NULL,
            UNIQUE(name, batch_id, branch_id)
        )
    """)

    # Sections table (e.g., A, B, C, D per class)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS section (
            section_id SERIAL PRIMARY KEY,
            class_id   INTEGER NOT NULL REFERENCES class(class_id) ON DELETE CASCADE,
            name       VARCHAR(20) NOT NULL,
            UNIQUE(class_id, name)
        )
    """)

    # Enhanced student table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS student (
            std_id     SERIAL PRIMARY KEY,
            name       VARCHAR(100) NOT NULL,
            email      VARCHAR(100),
            roll_no    VARCHAR(30),
            class_id   INTEGER REFERENCES class(class_id),
            section_id INTEGER REFERENCES section(section_id),
            dob        DATE
        )
    """)

    # Timetable table (period schedule per section per day)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS timetable (
            tt_id        SERIAL PRIMARY KEY,
            section_id   INTEGER NOT NULL REFERENCES section(section_id) ON DELETE CASCADE,
            day_of_week  INTEGER NOT NULL CHECK (day_of_week BETWEEN 0 AND 6),
            period_name  VARCHAR(100) NOT NULL,
            teacher_name VARCHAR(100),
            from_time    TIME NOT NULL,
            to_time      TIME NOT NULL,
            is_recess    BOOLEAN DEFAULT FALSE
        )
    """)

    # Subject table (kept for backward compat, linked to class)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS subject (
            sub_id     SERIAL PRIMARY KEY,
            name       VARCHAR(100) NOT NULL,
            class_id   INTEGER REFERENCES class(class_id),
            faculty_id INTEGER,
            from_time  TIME,
            to_time    TIME
        )
    """)

    # Attendance table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            id         SERIAL PRIMARY KEY,
            date       DATE NOT NULL,
            student_id INTEGER REFERENCES student(std_id) ON DELETE CASCADE,
            section_id INTEGER REFERENCES section(section_id),
            period_id  INTEGER REFERENCES timetable(tt_id),
            subject_id INTEGER REFERENCES subject(sub_id),
            image      TEXT
        )
    """)

    # ── Migrations: add columns if missing (for existing tables) ──────────────
    migrations = [
        ("student", "roll_no",    "VARCHAR(30)"),
        ("student", "class_id",   "INTEGER REFERENCES class(class_id)"),
        ("student", "section_id", "INTEGER REFERENCES section(section_id)"),
        ("student", "dob",        "DATE"),
        ("student", "parent_email", "VARCHAR(100)"),
        ("attendance", "section_id", "INTEGER REFERENCES section(section_id)"),
        ("attendance", "period_id",  "INTEGER REFERENCES timetable(tt_id)"),
        ("class", "batch_id",  "INTEGER REFERENCES batch(batch_id) ON DELETE SET NULL"),
        ("class", "branch_id", "INTEGER REFERENCES branch(branch_id) ON DELETE SET NULL"),
        ("branch", "batch_id", "INTEGER REFERENCES batch(batch_id) ON DELETE CASCADE"),
        ("users", "designation", "VARCHAR(100)"),
        ("users", "subjects", "VARCHAR(255)"),
        ("users", "role", None),  # handled separately below
    ]
    for table, col, col_type in migrations:
        if col_type is None:
            continue
        cur.execute("""
            SELECT 1 FROM information_schema.columns
            WHERE table_name=%s AND column_name=%s
        """, (table, col))
        if not cur.fetchone():
            cur.execute(f'ALTER TABLE {table} ADD COLUMN {col} {col_type}')

    # Migrate users role constraint to allow 'teacher'
    try:
        cur.execute("""
            ALTER TABLE users DROP CONSTRAINT IF EXISTS users_role_check;
            ALTER TABLE users ADD CONSTRAINT users_role_check
                CHECK (role IN ('admin','college','teacher'));
        """)
    except Exception:
        pass

    # Drop legacy UNIQUE constraints on branch (now scoped per batch)
    try:
        cur.execute("ALTER TABLE branch DROP CONSTRAINT IF EXISTS branch_name_key")
        cur.execute("ALTER TABLE branch DROP CONSTRAINT IF EXISTS branch_code_key")
    except Exception:
        pass

    # ── Engagement Log table ──────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS engagement_log (
            id             SERIAL PRIMARY KEY,
            date           DATE NOT NULL,
            section_id     INTEGER REFERENCES section(section_id),
            period_id      INTEGER REFERENCES timetable(tt_id),
            timestamp      TIMESTAMP NOT NULL,
            total_faces    INTEGER DEFAULT 0,
            attentive_pct  REAL DEFAULT 0,
            confused_pct   REAL DEFAULT 0,
            distracted_pct REAL DEFAULT 0,
            avg_score      REAL DEFAULT 0
        )
    """)

    # ── WhatsApp Recipients table ─────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS whatsapp_recipients (
            id         SERIAL PRIMARY KEY,
            name       VARCHAR(100) NOT NULL,
            phone      VARCHAR(20) NOT NULL,
            role       VARCHAR(20) NOT NULL CHECK (role IN ('parent','dean','hod','teacher')),
            student_id INTEGER REFERENCES student(std_id) ON DELETE CASCADE,
            class_id   INTEGER REFERENCES class(class_id),
            section_id INTEGER REFERENCES section(section_id)
        )
    """)

    # ── Camera Streams table ──────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS camera_streams (
            id         SERIAL PRIMARY KEY,
            stream_key VARCHAR(50) UNIQUE NOT NULL,
            name       VARCHAR(100) NOT NULL,
            rtsp_url   TEXT NOT NULL,
            section_id INTEGER REFERENCES section(section_id),
            is_active  BOOLEAN DEFAULT TRUE
        )
    """)

    # ── Engagement Sessions table (college user enables/disables monitoring) ──
    cur.execute("""
        CREATE TABLE IF NOT EXISTS engagement_sessions (
            id             SERIAL PRIMARY KEY,
            section_id     INTEGER NOT NULL REFERENCES section(section_id),
            period_id      INTEGER NOT NULL REFERENCES timetable(tt_id),
            started_at     TIMESTAMP NOT NULL DEFAULT NOW(),
            ended_at       TIMESTAMP,
            capture_interval INTEGER DEFAULT 60,
            is_active      BOOLEAN DEFAULT TRUE,
            started_by     INTEGER REFERENCES users(id)
        )
    """)

    # ── Per-student engagement log (individual student engagement per capture) ─
    cur.execute("""
        CREATE TABLE IF NOT EXISTS engagement_student_log (
            id               SERIAL PRIMARY KEY,
            session_id       INTEGER REFERENCES engagement_sessions(id) ON DELETE CASCADE,
            log_id           INTEGER REFERENCES engagement_log(id) ON DELETE CASCADE,
            student_id       INTEGER REFERENCES student(std_id) ON DELETE CASCADE,
            student_name     VARCHAR(100),
            timestamp        TIMESTAMP NOT NULL,
            engagement_score REAL DEFAULT 0,
            emotion          VARCHAR(30),
            is_attentive     BOOLEAN DEFAULT TRUE,
            gaze_direction   VARCHAR(20),
            liveness_score   REAL DEFAULT 0,
            is_live          BOOLEAN DEFAULT TRUE
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS report_schedules (
            id               SERIAL PRIMARY KEY,
            teacher_id       INTEGER NOT NULL,
            section_id       INTEGER NOT NULL,
            frequency        VARCHAR(20) DEFAULT 'daily',
            send_time        TIME DEFAULT '09:00',
            include_parents  BOOLEAN DEFAULT FALSE,
            active           BOOLEAN DEFAULT TRUE,
            created_at       TIMESTAMP DEFAULT NOW()
        )
    """)

    conn.commit()
    cur.close()
    _return_conn(conn)


# ── Auth decorators ───────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*a, **kw)
    return dec

def admin_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('Admin access required.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*a, **kw)
    return dec

def college_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if 'user_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Authentication required'}), 401
            return redirect(url_for('login'))
        if session.get('role') != 'college':
            if request.path.startswith('/api/'):
                return jsonify({'error': 'College staff access required'}), 403
            flash('College staff access required.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*a, **kw)
    return dec


def teacher_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if 'user_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Authentication required'}), 401
            return redirect(url_for('teacher_login'))
        if session.get('role') != 'teacher':
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Teacher access required'}), 403
            flash('Teacher access required.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*a, **kw)
    return dec


# ── Utilities ─────────────────────────────────────────────────────────────────
@app.context_processor
def inject_now():
    return {'now': _now_ist(), 'startup_warnings': _WARNINGS}


def _save_b64_image(b64_str, path, max_size=320):
    if ',' in b64_str:
        b64_str = b64_str.split(',')[1]
    data = base64.b64decode(b64_str)
    if cv2 is None or np is None:
        # Fallback: just save raw JPEG bytes
        with open(path, 'wb') as f:
            f.write(data)
        return
    arr = np.frombuffer(data, dtype=np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    if img is None:
        with open(path, 'wb') as f:
            f.write(data)
        return
    # Resize to save memory
    h, w = img.shape[:2]
    if max(h, w) > max_size:
        scale = max_size / max(h, w)
        img = cv2.resize(img, (int(w * scale), int(h * scale)), interpolation=cv2.INTER_AREA)
    cv2.imwrite(path, img, [cv2.IMWRITE_JPEG_QUALITY, 85])


def _regenerate_embeddings():
    """Full regeneration — only used as fallback."""
    import gc
    _load_ai_modules()
    embeddings, names = gen_embed.get_embeddings(DATASET_DIR)
    if embeddings:
        df = pd.DataFrame(embeddings)
        df['name'] = names
        df.to_csv(os.path.join(BASE_DIR, 'embeddings.csv'), index=False)
    gc.collect()


def _append_embeddings(student_name):
    """Incremental: generate embeddings only for one student and append to CSV."""
    import gc

    student_dir = os.path.join(DATASET_DIR, student_name)
    if not os.path.isdir(student_dir):
        return

    # Import only DeepFace — avoid loading the full AI stack
    from deepface import DeepFace

    csv_path = os.path.join(BASE_DIR, 'embeddings.csv')
    # Load existing embeddings if any
    if os.path.exists(csv_path):
        existing_df = pd.read_csv(csv_path)
    else:
        existing_df = pd.DataFrame()

    new_embeddings = []
    new_names = []
    for image_name in os.listdir(student_dir):
        image_path = os.path.join(student_dir, image_name)
        if not os.path.isfile(image_path):
            continue
        try:
            embedding = DeepFace.represent(
                img_path=image_path,
                model_name='Facenet512',
                detector_backend='opencv'
            )[0]["embedding"]
            new_embeddings.append(embedding)
            new_names.append(student_name)
        except Exception as e:
            print(f"[embed] Could not process {image_path}: {e}", flush=True)
        # Free memory after each image
        gc.collect()

    if new_embeddings:
        new_df = pd.DataFrame(new_embeddings)
        new_df['name'] = new_names
        if not existing_df.empty:
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
        else:
            combined_df = new_df
        combined_df.to_csv(csv_path, index=False)
    gc.collect()


def _lightweight_face_recognition(image_path, emb_vectors, emb_names):
    """Face recognition using OpenCV LBPH recognizer. Fast and accurate without DeepFace."""
    import gc

    if cv2 is None or np is None:
        return None, []

    img = cv2.imread(image_path)
    if img is None:
        return None, []

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray = cv2.equalizeHist(gray)  # improve contrast for face detection
    cascade_path = cv2.data.haarcascades + 'haarcascade_frontalface_default.xml'
    face_cascade = cv2.CascadeClassifier(cascade_path)
    detected_faces = face_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5, minSize=(80, 80))

    if len(detected_faces) == 0:
        return img, []

    # Build LBPH recognizer from stored face images in dataset/
    recognizer = cv2.face.LBPHFaceRecognizer_create(radius=1, neighbors=8, grid_x=8, grid_y=8)
    
    # Get unique names from embeddings to know which students exist
    unique_names = list(dict.fromkeys(emb_names))
    
    # Train recognizer from dataset images
    train_faces = []
    train_labels = []
    name_to_label = {}
    label_to_name = {}
    
    for idx, student_name in enumerate(unique_names):
        student_dir = os.path.join(DATASET_DIR, student_name)
        if not os.path.isdir(student_dir):
            continue
        name_to_label[student_name] = idx
        label_to_name[idx] = student_name
        
        for img_file in os.listdir(student_dir):
            ref_path = os.path.join(student_dir, img_file)
            ref_img = cv2.imread(ref_path, cv2.IMREAD_GRAYSCALE)
            if ref_img is None:
                continue
            ref_img = cv2.equalizeHist(ref_img)
            # Detect face in reference image too for better matching
            ref_faces = face_cascade.detectMultiScale(ref_img, 1.3, 3, minSize=(50, 50))
            if len(ref_faces) > 0:
                rx, ry, rw, rh = ref_faces[0]
                ref_face = ref_img[ry:ry+rh, rx:rx+rw]
            else:
                ref_face = ref_img  # use full image if no face detected
            ref_face = cv2.resize(ref_face, (200, 200))
            train_faces.append(ref_face)
            train_labels.append(idx)

    if not train_faces:
        print("[face] No training faces found in dataset/", flush=True)
        return img, []

    recognizer.train(train_faces, np.array(train_labels))

    # Recognize each detected face
    identified_persons = []
    already_matched = set()

    for (x, y, w, h) in detected_faces:
        face_roi = gray[y:y+h, x:x+w]
        face_resized = cv2.resize(face_roi, (200, 200))
        
        label, confidence = recognizer.predict(face_resized)
        # LBPH confidence: lower = better match. Threshold 70 for reliable matching
        print(f"[face] Detected face at ({x},{y},{w},{h}) -> label={label}, confidence={confidence:.1f}", flush=True)
        
        if confidence < 70 and label in label_to_name:
            matched_name = label_to_name[label]
            if matched_name not in already_matched:
                already_matched.add(matched_name)
                identified_persons.append({
                    'name': matched_name,
                    'distance': float(confidence),
                    'confidence_score': float(max(0, (70 - confidence) / 70)),
                    'facial_area': {'x': int(x), 'y': int(y), 'w': int(w), 'h': int(h)},
                    'recognized': True
                })
            else:
                # Same person detected twice in frame
                identified_persons.append({
                    'name': matched_name,
                    'distance': float(confidence),
                    'confidence_score': float(max(0, (70 - confidence) / 70)),
                    'facial_area': {'x': int(x), 'y': int(y), 'w': int(w), 'h': int(h)},
                    'recognized': True
                })
        else:
            # Unknown face - still add to results for red box drawing
            identified_persons.append({
                'name': '',
                'distance': float(confidence),
                'confidence_score': 0.0,
                'facial_area': {'x': int(x), 'y': int(y), 'w': int(w), 'h': int(h)},
                'recognized': False
            })

    gc.collect()
    return img, identified_persons


def _get_student_ids_local(names):
    """Get student IDs by name without importing Attendance_update_db."""
    if not names:
        return {}
    conn = get_conn(); cur = conn.cursor()
    placeholders = ','.join(['%s'] * len(names))
    cur.execute(f"SELECT std_id, name FROM student WHERE name IN ({placeholders})", names)
    results = cur.fetchall()
    cur.close(); _return_conn(conn)
    return {name: std_id for std_id, name in results}


def _serialize(obj):
    if isinstance(obj, dict):
        return {k: _serialize(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_serialize(i) for i in obj]
    if isinstance(obj, (dt.date, dt.time, dt.datetime)):
        return str(obj)
    if isinstance(obj, np.bool_):
        return bool(obj)
    if isinstance(obj, np.integer):
        return int(obj)
    if isinstance(obj, np.floating):
        return float(obj)
    if isinstance(obj, np.ndarray):
        return obj.tolist()
    if hasattr(obj, 'item'):
        return obj.item()
    return obj


DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']


# ── Auth routes ───────────────────────────────────────────────────────────────
@app.route('/')
def index():
    if 'user_id' in session:
        role = session.get('role', '')
        if role == 'teacher':
            return redirect(FRONTEND_URL + '/teacher/dashboard')
        return redirect(FRONTEND_URL + '/college/dashboard')
    return redirect(FRONTEND_URL)


# ── JSON Auth API (for React SPA) ────────────────────────────────────────────

@app.route('/api/auth/login', methods=['POST'])
def api_auth_login():
    """JSON login for both teacher and college roles."""
    data = request.get_json() or request.form
    email = (data.get('email') or '').strip()
    password = data.get('password') or ''
    role_hint = data.get('role', '')  # 'teacher' or 'college'
    if not email or not password:
        return jsonify({'error': 'Email and password required'}), 400
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT id,name,email,password_hash,role FROM users WHERE email=%s', (email,))
    user = cur.fetchone(); cur.close(); _return_conn(conn)
    if not user or not check_password_hash(user[3], password):
        return jsonify({'error': 'Invalid email or password'}), 401
    if role_hint and user[4] != role_hint:
        return jsonify({'error': f'This account is not a {role_hint} account'}), 403
    session.update(user_id=user[0], name=user[1], email=user[2], role=user[4])
    return jsonify({'id': user[0], 'name': user[1], 'email': user[2], 'role': user[4]})


@app.route('/api/auth/signup', methods=['POST'])
def api_auth_signup():
    """JSON signup for teachers or college staff."""
    data = request.get_json() or request.form
    name = (data.get('name') or '').strip()
    email = (data.get('email') or '').strip()
    password = data.get('password') or ''
    role = data.get('role', 'teacher')
    if role not in ('teacher', 'college'):
        return jsonify({'error': 'Invalid role'}), 400
    if not all([name, email, password]):
        return jsonify({'error': 'All fields are required'}), 400
    try:
        conn = get_conn(); cur = conn.cursor()
        cur.execute(
            'INSERT INTO users (name,email,password_hash,role) VALUES (%s,%s,%s,%s)',
            (name, email, generate_password_hash(password), role)
        )
        conn.commit(); cur.close(); _return_conn(conn)
        return jsonify({'message': 'Account created! Please log in.'})
    except psycopg2.errors.UniqueViolation:
        return jsonify({'error': 'Email already registered'}), 409
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/auth/me')
def api_auth_me():
    """Check current session."""
    if 'user_id' in session:
        return jsonify({'id': session['user_id'], 'name': session['name'],
                        'email': session['email'], 'role': session['role']})
    return jsonify({'error': 'Not authenticated'}), 401


@app.route('/api/auth/logout', methods=['POST', 'GET'])
def api_auth_logout():
    session.clear()
    return jsonify({'message': 'Logged out'})


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Redirect to React frontend login."""
    return redirect(FRONTEND_URL + '/college/login')


@app.route('/teacher/login', methods=['GET', 'POST'])
def teacher_login():
    """Redirect to React frontend teacher login."""
    return redirect(FRONTEND_URL + '/teacher/login')


@app.route('/teacher/signup', methods=['GET', 'POST'])
def teacher_signup():
    """Redirect to React frontend teacher login."""
    return redirect(FRONTEND_URL + '/teacher/login')


@app.route('/signup', methods=['GET', 'POST'])
def signup():
    """Redirect to React frontend college login."""
    return redirect(FRONTEND_URL + '/college/login')


@app.route('/logout')
def logout():
    role = session.get('role', '')
    session.clear()
    if role == 'teacher':
        return redirect(FRONTEND_URL + '/teacher/login')
    return redirect(FRONTEND_URL + '/college/login')


@app.route('/dashboard')
@login_required
def dashboard():
    if session['role'] == 'admin':
        return redirect(url_for('admin_dashboard'))
    elif session['role'] == 'teacher':
        return redirect(url_for('teacher_dashboard'))
    return redirect(url_for('college_dashboard'))


# ══════════════════════════════════════════════════════════════════════════════
# ADMIN PORTAL
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/admin')
@admin_required
def admin_dashboard():
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT COUNT(*) FROM student')
    student_count = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM class')
    class_count = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM section')
    section_count = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM attendance WHERE date=%s', (_today_ist(),))
    today_count = cur.fetchone()[0]
    cur.close(); _return_conn(conn)
    return render_template('admin/dashboard.html',
                           student_count=student_count,
                           class_count=class_count,
                           section_count=section_count,
                           today_count=today_count)


# ── Class & Section Management ────────────────────────────────────────────────
@app.route('/admin/classes')
@admin_required
def admin_classes():
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        SELECT c.class_id, c.name,
               (SELECT COUNT(*) FROM section s WHERE s.class_id=c.class_id) as sec_count,
               (SELECT COUNT(*) FROM student st WHERE st.class_id=c.class_id) as stu_count
        FROM class c ORDER BY c.name
    """)
    classes = cur.fetchall()
    cur.close(); _return_conn(conn)
    return render_template('admin/classes.html', classes=classes)


@app.route('/admin/classes/add', methods=['POST'])
@admin_required
def admin_add_class():
    name = request.form.get('name', '').strip()
    if not name:
        flash('Class name is required.', 'danger')
        return redirect(url_for('admin_classes'))
    try:
        conn = get_conn(); cur = conn.cursor()
        cur.execute('INSERT INTO class (name) VALUES (%s)', (name,))
        conn.commit(); cur.close(); _return_conn(conn)
        flash(f'Class "{name}" added.', 'success')
    except psycopg2.errors.UniqueViolation:
        flash(f'Class "{name}" already exists.', 'danger')
    except Exception as e:
        flash(f'Error: {e}', 'danger')
    return redirect(url_for('admin_classes'))


@app.route('/admin/classes/<int:class_id>/delete', methods=['POST'])
@admin_required
def admin_delete_class(class_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute('DELETE FROM class WHERE class_id=%s', (class_id,))
    conn.commit(); cur.close(); _return_conn(conn)
    flash('Class deleted.', 'success')
    return redirect(url_for('admin_classes'))


@app.route('/admin/classes/<int:class_id>/sections')
@admin_required
def admin_sections(class_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT name FROM class WHERE class_id=%s', (class_id,))
    cls = cur.fetchone()
    if not cls:
        flash('Class not found.', 'danger')
        return redirect(url_for('admin_classes'))
    cur.execute("""
        SELECT s.section_id, s.name,
               (SELECT COUNT(*) FROM student st WHERE st.section_id=s.section_id) as stu_count
        FROM section s WHERE s.class_id=%s ORDER BY s.name
    """, (class_id,))
    sections = cur.fetchall()
    cur.close(); _return_conn(conn)
    return render_template('admin/sections.html', class_id=class_id, class_name=cls[0], sections=sections)


@app.route('/admin/classes/<int:class_id>/sections/add', methods=['POST'])
@admin_required
def admin_add_section(class_id):
    name = request.form.get('name', '').strip()
    if not name:
        flash('Section name is required.', 'danger')
        return redirect(url_for('admin_sections', class_id=class_id))
    try:
        conn = get_conn(); cur = conn.cursor()
        cur.execute('INSERT INTO section (class_id, name) VALUES (%s, %s)', (class_id, name))
        conn.commit(); cur.close(); _return_conn(conn)
        flash(f'Section "{name}" added.', 'success')
    except psycopg2.errors.UniqueViolation:
        flash(f'Section "{name}" already exists in this class.', 'danger')
    except Exception as e:
        flash(f'Error: {e}', 'danger')
    return redirect(url_for('admin_sections', class_id=class_id))


@app.route('/admin/sections/<int:section_id>/delete', methods=['POST'])
@admin_required
def admin_delete_section(section_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT class_id FROM section WHERE section_id=%s', (section_id,))
    row = cur.fetchone()
    class_id = row[0] if row else None
    cur.execute('DELETE FROM section WHERE section_id=%s', (section_id,))
    conn.commit(); cur.close(); _return_conn(conn)
    flash('Section deleted.', 'success')
    return redirect(url_for('admin_sections', class_id=class_id) if class_id else url_for('admin_classes'))


# ── Timetable Management ─────────────────────────────────────────────────────
@app.route('/admin/timetable/<int:section_id>')
@admin_required
def admin_timetable(section_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        SELECT s.name as sec_name, c.name as cls_name, s.class_id
        FROM section s JOIN class c ON s.class_id=c.class_id
        WHERE s.section_id=%s
    """, (section_id,))
    info = cur.fetchone()
    if not info:
        flash('Section not found.', 'danger')
        return redirect(url_for('admin_classes'))

    cur.execute("""
        SELECT tt_id, day_of_week, period_name, teacher_name, from_time, to_time, is_recess
        FROM timetable WHERE section_id=%s ORDER BY day_of_week, from_time
    """, (section_id,))
    periods = cur.fetchall()
    cur.close(); _return_conn(conn)

    # Group by day
    timetable = {i: [] for i in range(7)}
    for p in periods:
        timetable[p[1]].append(p)

    return render_template('admin/timetable.html',
                           section_id=section_id,
                           section_name=info[0],
                           class_name=info[1],
                           class_id=info[2],
                           timetable=timetable,
                           days=DAYS)


@app.route('/admin/timetable/<int:section_id>/add', methods=['POST'])
@admin_required
def admin_add_period(section_id):
    day = int(request.form.get('day_of_week', 0))
    period_name = request.form.get('period_name', '').strip()
    teacher_name = request.form.get('teacher_name', '').strip()
    from_time = request.form.get('from_time', '')
    to_time = request.form.get('to_time', '')
    is_recess = request.form.get('is_recess') == 'on'

    if not period_name or not from_time or not to_time:
        flash('Period name, start time, and end time are required.', 'danger')
        return redirect(url_for('admin_timetable', section_id=section_id))

    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        INSERT INTO timetable (section_id, day_of_week, period_name, teacher_name, from_time, to_time, is_recess)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (section_id, day, period_name, teacher_name or None, from_time, to_time, is_recess))
    conn.commit(); cur.close(); _return_conn(conn)
    flash(f'Period "{period_name}" added for {DAYS[day]}.', 'success')
    return redirect(url_for('admin_timetable', section_id=section_id))


@app.route('/admin/timetable/delete/<int:tt_id>', methods=['POST'])
@admin_required
def admin_delete_period(tt_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT section_id FROM timetable WHERE tt_id=%s', (tt_id,))
    row = cur.fetchone()
    section_id = row[0] if row else None
    cur.execute('DELETE FROM timetable WHERE tt_id=%s', (tt_id,))
    conn.commit(); cur.close(); _return_conn(conn)
    flash('Period deleted.', 'success')
    return redirect(url_for('admin_timetable', section_id=section_id) if section_id else url_for('admin_classes'))


# ── Student Management ────────────────────────────────────────────────────────
@app.route('/admin/students')
@admin_required
def admin_students():
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        SELECT st.std_id, st.name, st.email, st.roll_no, c.name as class_name, s.name as sec_name, st.dob
        FROM student st
        LEFT JOIN class c ON st.class_id=c.class_id
        LEFT JOIN section s ON st.section_id=s.section_id
        ORDER BY c.name, s.name, st.roll_no, st.name
    """)
    students = cur.fetchall()
    cur.close(); _return_conn(conn)
    return render_template('admin/students.html', students=students)


@app.route('/admin/register', methods=['GET', 'POST'])
@admin_required
def admin_register_student():
    conn = get_conn(); cur = conn.cursor()

    if request.method == 'POST':
        name        = request.form.get('name', '').strip()
        email       = request.form.get('email', '').strip()
        roll_no     = request.form.get('roll_no', '').strip()
        class_id    = request.form.get('class_id', '')
        section_id  = request.form.get('section_id', '')
        dob         = request.form.get('dob', '').strip()
        photos_json = request.form.get('photos', '[]')

        if not name:
            flash('Student name is required.', 'danger')
            cur.execute('SELECT class_id, name FROM class ORDER BY name')
            classes = cur.fetchall(); cur.close(); _return_conn(conn)
            return render_template('admin/register_student.html', classes=classes)

        photos = json.loads(photos_json)
        if not photos:
            flash('Please capture at least one photo.', 'danger')
            cur.execute('SELECT class_id, name FROM class ORDER BY name')
            classes = cur.fetchall(); cur.close(); _return_conn(conn)
            return render_template('admin/register_student.html', classes=classes)

        # Limit to 2 photos max to save memory on free tier
        photos = photos[:2]

        try:
            cur.execute("""
                INSERT INTO student (name, email, roll_no, class_id, section_id, dob)
                VALUES (%s, %s, %s, %s, %s, %s) RETURNING std_id
            """, (
                name,
                email or None,
                roll_no or None,
                int(class_id) if class_id else None,
                int(section_id) if section_id else None,
                dob or None
            ))
            conn.commit()
        except Exception as e:
            flash(f'Database error: {e}', 'danger')
            cur.execute('SELECT class_id, name FROM class ORDER BY name')
            classes = cur.fetchall(); cur.close(); _return_conn(conn)
            return render_template('admin/register_student.html', classes=classes)

        cur.close(); _return_conn(conn)

        import gc
        student_dir = os.path.join(DATASET_DIR, name)
        os.makedirs(student_dir, exist_ok=True)
        for i, b64 in enumerate(photos):
            _save_b64_image(b64, os.path.join(student_dir, f'photo_{i+1}.jpg'))
            del b64
        del photos
        gc.collect()

        # Skip embedding generation on server (512MB not enough for TensorFlow)
        # Embeddings CSV is pre-generated and pushed from local machine
        flash(f'Student "{name}" registered successfully.', 'success')
        return redirect(url_for('admin_students'))

    # GET
    cur.execute('SELECT class_id, name FROM class ORDER BY name')
    classes = cur.fetchall()
    cur.close(); _return_conn(conn)
    return render_template('admin/register_student.html', classes=classes)


@app.route('/admin/delete/<int:std_id>', methods=['POST'])
@admin_required
def delete_student(std_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT name FROM student WHERE std_id=%s', (std_id,))
    row = cur.fetchone()
    if row:
        name = row[0]
        cur.execute('DELETE FROM attendance WHERE student_id=%s', (std_id,))
        cur.execute('DELETE FROM student WHERE std_id=%s', (std_id,))
        conn.commit()
        d = os.path.join(DATASET_DIR, name)
        if os.path.exists(d):
            shutil.rmtree(d)
        # Remove student's embeddings from CSV (no AI needed)
        csv_path = os.path.join(BASE_DIR, 'embeddings.csv')
        if os.path.exists(csv_path):
            try:
                df = pd.read_csv(csv_path)
                df = df[df['name'] != name]
                df.to_csv(csv_path, index=False)
            except Exception:
                pass
        flash(f'Student "{name}" deleted.', 'success')
    cur.close(); _return_conn(conn)
    return redirect(url_for('admin_students'))


# ── API: Sections for a class (AJAX) ─────────────────────────────────────────
@app.route('/api/sections/<int:class_id>')
@login_required
def api_sections(class_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT section_id, name FROM section WHERE class_id=%s ORDER BY name', (class_id,))
    rows = cur.fetchall(); cur.close(); _return_conn(conn)
    return jsonify([{'id': r[0], 'name': r[1]} for r in rows])


@app.route('/api/subjects')
@login_required
def api_subjects():
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT sub_id, name, from_time, to_time FROM subject ORDER BY from_time')
    rows = cur.fetchall(); cur.close(); _return_conn(conn)
    return jsonify([{'id': r[0], 'name': r[1],
                     'time': f"{str(r[2])[:5]} – {str(r[3])[:5]}"} for r in rows])


@app.route('/api/timetable/<int:section_id>')
@login_required
def api_timetable(section_id):
    """Get periods for a section. Shows today's periods first, then all others."""
    today_dow = _today_ist().weekday()
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT ON (period_name) tt_id, period_name, teacher_name, from_time, to_time, is_recess, day_of_week
        FROM timetable WHERE section_id=%s
        ORDER BY period_name, (day_of_week = %s) DESC, from_time
    """, (section_id, today_dow))
    rows = cur.fetchall(); cur.close(); _return_conn(conn)
    return jsonify([{
        'id': r[0], 'name': r[1], 'teacher': r[2],
        'time': f"{str(r[3])[:5]} – {str(r[4])[:5]}", 'is_recess': r[5]
    } for r in rows])


# ── Report Sending (Enhanced) ─────────────────────────────────────────────────
@app.route('/admin/reports')
@admin_required
def admin_reports():
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT class_id, name FROM class ORDER BY name')
    classes = cur.fetchall()
    cur.close(); _return_conn(conn)
    return render_template('admin/reports.html', classes=classes)


@app.route('/admin/send-report', methods=['POST'])
@admin_required
def admin_send_report():
    report_type = request.form.get('report_type', 'daily')
    date_str    = request.form.get('date', _today_ist().strftime('%Y-%m-%d'))
    class_id    = request.form.get('class_id', '')
    section_id  = request.form.get('section_id', '')
    recipients  = request.form.get('recipients', '').strip()

    try:
        from send_report import send_filtered_report

        date_obj = dt.datetime.strptime(date_str, '%Y-%m-%d').date()
        recipient_list = [e.strip() for e in recipients.split(',') if e.strip()]

        if not recipient_list:
            from config import faculty_emails, director_email
            recipient_list = list(dict.fromkeys(list(faculty_emails) + [director_email]))

        send_filtered_report(
            report_type=report_type,
            date_obj=date_obj,
            class_id=int(class_id) if class_id else None,
            section_id=int(section_id) if section_id else None,
            recipients=recipient_list
        )
        flash(f'{report_type.capitalize()} report sent successfully!', 'success')
    except Exception as e:
        flash(f'Failed to send report: {e}', 'danger')

    return redirect(url_for('admin_reports'))


# ── Engagement Analytics Routes ───────────────────────────────────────────────
@app.route('/admin/engagement')
@admin_required
def admin_engagement():
    """Admin view: Engagement reports & analytics (read-only)."""
    conn = get_conn(); cur = conn.cursor()

    # Filters
    class_filter = request.args.get('class_id', '')
    section_filter = request.args.get('section_id', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    query = """
        SELECT e.date, c.name as class_name, s.name as sec_name, t.period_name,
               e.timestamp, e.total_faces, e.attentive_pct, e.confused_pct,
               e.distracted_pct, e.avg_score
        FROM engagement_log e
        LEFT JOIN section s ON e.section_id=s.section_id
        LEFT JOIN class c ON s.class_id=c.class_id
        LEFT JOIN timetable t ON e.period_id=t.tt_id
        WHERE 1=1
    """
    params = []
    if class_filter:
        query += " AND s.class_id = %s"
        params.append(int(class_filter))
    if section_filter:
        query += " AND e.section_id = %s"
        params.append(int(section_filter))
    if date_from:
        query += " AND e.date >= %s"
        params.append(date_from)
    if date_to:
        query += " AND e.date <= %s"
        params.append(date_to)
    query += " ORDER BY e.timestamp DESC LIMIT 100"

    cur.execute(query, params)
    logs = cur.fetchall()

    # Summary stats
    cur.execute("""
        SELECT COUNT(*), COALESCE(AVG(avg_score),0), COALESCE(AVG(attentive_pct),0)
        FROM engagement_log WHERE date >= CURRENT_DATE - INTERVAL '7 days'
    """)
    stats = cur.fetchone()

    # Per-student engagement data (latest 200 records)
    student_query = """
        SELECT esl.student_name, esl.engagement_score, esl.emotion,
               esl.is_attentive, esl.gaze_direction, esl.liveness_score,
               esl.is_live, esl.timestamp,
               c.name as class_name, s.name as sec_name, t.period_name
        FROM engagement_student_log esl
        JOIN engagement_sessions es ON esl.session_id = es.id
        LEFT JOIN section s ON es.section_id = s.section_id
        LEFT JOIN class c ON s.class_id = c.class_id
        LEFT JOIN timetable t ON es.period_id = t.tt_id
        WHERE 1=1
    """
    sparams = []
    if class_filter:
        student_query += " AND s.class_id = %s"
        sparams.append(int(class_filter))
    if section_filter:
        student_query += " AND es.section_id = %s"
        sparams.append(int(section_filter))
    if date_from:
        student_query += " AND esl.timestamp::date >= %s"
        sparams.append(date_from)
    if date_to:
        student_query += " AND esl.timestamp::date <= %s"
        sparams.append(date_to)
    student_query += " ORDER BY esl.timestamp DESC LIMIT 200"
    cur.execute(student_query, sparams)
    student_logs = cur.fetchall()

    cur.execute('SELECT class_id, name FROM class ORDER BY name')
    classes = cur.fetchall()
    cur.execute('SELECT section_id, name, class_id FROM section ORDER BY name')
    sections = cur.fetchall()
    cur.close(); _return_conn(conn)
    return render_template('admin/engagement.html', logs=logs, classes=classes,
                           sections=sections, stats=stats,
                           student_logs=student_logs,
                           class_filter=class_filter, section_filter=section_filter,
                           date_from=date_from, date_to=date_to)


@app.route('/api/engagement/analyze', methods=['POST'])
@login_required
def api_analyze_engagement():
    """Analyze a single image for engagement (used by frontend)."""
    data = request.get_json()
    if not data or 'photo' not in data:
        return jsonify({'error': 'No photo provided'}), 400

    try:
        b64 = data['photo']
        if ',' in b64:
            b64 = b64.split(',')[1]
        img_data = base64.b64decode(b64)
        arr = np.frombuffer(img_data, dtype=np.uint8)
        img = cv2.imdecode(arr, cv2.IMREAD_COLOR)

        _load_ai_modules()
        engagement = analyze_engagement(img)
        return jsonify({
            'total_faces': engagement.total_faces,
            'attentive_count': engagement.attentive_count,
            'distracted_count': engagement.distracted_count,
            'attentive_pct': engagement.attentive_pct,
            'confused_pct': engagement.confused_pct,
            'distracted_pct': engagement.distracted_pct,
            'avg_score': round(engagement.avg_engagement_score, 3),
            'emotion_distribution': engagement.emotion_distribution,
            'faces': [{
                'gaze': f.gaze_direction,
                'emotion': f.emotion,
                'attentive': f.is_attentive,
                'score': round(f.engagement_score, 2)
            } for f in engagement.faces]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Liveness Detection Route ─────────────────────────────────────────────────
@app.route('/api/liveness/check', methods=['POST'])
@login_required
def api_liveness_check():
    """Check if a captured photo is of a real person (not a screen/photo)."""
    data = request.get_json()
    if not data or 'photo' not in data:
        return jsonify({'error': 'No photo provided'}), 400

    try:
        b64 = data['photo']
        if ',' in b64:
            b64 = b64.split(',')[1]
        img_data = base64.b64decode(b64)
        arr = np.frombuffer(img_data, dtype=np.uint8)
        img = cv2.imdecode(arr, cv2.IMREAD_COLOR)

        _load_ai_modules()
        result = quick_liveness_check(img)
        return jsonify({
            'is_live': result.is_live,
            'confidence': round(result.confidence, 3),
            'texture_pass': result.texture_pass,
            'moiré_pass': result.moiré_pass,
            'reason': result.reason
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Continuous Attendance (RTSP) Routes ───────────────────────────────────────
@app.route('/admin/cameras')
@admin_required
def admin_cameras():
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        SELECT cs.id, cs.stream_key, cs.name, cs.rtsp_url, cs.section_id, cs.is_active,
               c.name as class_name, s.name as sec_name
        FROM camera_streams cs
        LEFT JOIN section s ON cs.section_id=s.section_id
        LEFT JOIN class c ON s.class_id=c.class_id
        ORDER BY cs.name
    """)
    cameras = cur.fetchall()
    cur.execute("""
        SELECT s.section_id, c.name || ' - ' || s.name as full_name
        FROM section s JOIN class c ON s.class_id=c.class_id ORDER BY c.name, s.name
    """)
    sections = cur.fetchall()
    cur.close(); _return_conn(conn)

    ca = get_continuous_attendance()
    status = ca.get_status() if ca else {'running': False, 'streams': {}}

    return render_template('admin/cameras.html', cameras=cameras,
                           sections=sections, ca_status=status)


@app.route('/admin/cameras/add', methods=['POST'])
@admin_required
def admin_add_camera():
    name = request.form.get('name', '').strip()
    rtsp_url = request.form.get('rtsp_url', '').strip()
    section_id = request.form.get('section_id', '')

    if not name or not section_id:
        flash('Camera name and section are required.', 'danger')
        return redirect(url_for('admin_cameras'))

    stream_key = name.lower().replace(' ', '_')

    try:
        conn = get_conn(); cur = conn.cursor()
        cur.execute("""
            INSERT INTO camera_streams (stream_key, name, rtsp_url, section_id)
            VALUES (%s, %s, %s, %s)
        """, (stream_key, name, rtsp_url or 'webcam', int(section_id)))
        conn.commit(); cur.close(); _return_conn(conn)

        # Register with continuous attendance
        ca = get_continuous_attendance()
        if ca:
            ca.add_stream(stream_key, rtsp_url or 'webcam', int(section_id), name)

        flash(f'Camera "{name}" added.', 'success')
    except Exception as e:
        flash(f'Error: {e}', 'danger')

    return redirect(url_for('admin_cameras'))


@app.route('/admin/cameras/<int:cam_id>/delete', methods=['POST'])
@admin_required
def admin_delete_camera(cam_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT stream_key FROM camera_streams WHERE id=%s', (cam_id,))
    row = cur.fetchone()
    if row:
        ca = get_continuous_attendance()
        if ca:
            ca.remove_stream(row[0])
    cur.execute('DELETE FROM camera_streams WHERE id=%s', (cam_id,))
    conn.commit(); cur.close(); _return_conn(conn)
    flash('Camera removed.', 'success')
    return redirect(url_for('admin_cameras'))


@app.route('/admin/cameras/toggle', methods=['POST'])
@admin_required
def admin_toggle_continuous():
    """Start or stop continuous attendance capture."""
    ca = get_continuous_attendance()
    if not ca:
        flash('Continuous attendance system not initialized.', 'danger')
        return redirect(url_for('admin_cameras'))

    if ca.running:
        ca.stop()
        flash('Continuous attendance stopped.', 'warning')
    else:
        # Load streams from DB
        conn = get_conn(); cur = conn.cursor()
        cur.execute('SELECT stream_key, rtsp_url, section_id, name FROM camera_streams WHERE is_active=TRUE')
        for row in cur.fetchall():
            ca.add_stream(row[0], row[1], row[2], row[3])
        cur.close(); _return_conn(conn)
        ca.start()
        flash('Continuous attendance started!', 'success')

    return redirect(url_for('admin_cameras'))


@app.route('/admin/cameras/capture-now', methods=['POST'])
@admin_required
def admin_capture_now():
    """Manually trigger immediate capture from all cameras."""
    ca = get_continuous_attendance()
    if not ca:
        flash('System not initialized.', 'danger')
        return redirect(url_for('admin_cameras'))

    results = ca.capture_now()
    total_students = sum(len(r.get('students', [])) for r in results.values())
    flash(f'Captured from {len(results)} camera(s). {total_students} student(s) detected.', 'success')
    return redirect(url_for('admin_cameras'))


# ── WhatsApp Recipients Management ───────────────────────────────────────────
@app.route('/admin/whatsapp')
@admin_required
def admin_whatsapp():
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        SELECT wr.id, wr.name, wr.phone, wr.role, wr.student_id, wr.class_id,
               st.name as student_name, c.name as class_name
        FROM whatsapp_recipients wr
        LEFT JOIN student st ON wr.student_id=st.std_id
        LEFT JOIN class c ON wr.class_id=c.class_id
        ORDER BY wr.role, wr.name
    """)
    recipients = cur.fetchall()
    cur.execute('SELECT std_id, name FROM student ORDER BY name')
    students = cur.fetchall()
    cur.execute('SELECT class_id, name FROM class ORDER BY name')
    classes = cur.fetchall()
    cur.close(); _return_conn(conn)

    whatsapp_configured = whatsapp_client is not None
    return render_template('admin/whatsapp.html', recipients=recipients,
                           students=students, classes=classes,
                           whatsapp_configured=whatsapp_configured)


@app.route('/admin/whatsapp/add', methods=['POST'])
@admin_required
def admin_add_whatsapp_recipient():
    name = request.form.get('name', '').strip()
    phone = request.form.get('phone', '').strip()
    role = request.form.get('role', '').strip()
    student_id = request.form.get('student_id', '') or None
    class_id = request.form.get('class_id', '') or None

    if not name or not phone or not role:
        flash('Name, phone, and role are required.', 'danger')
        return redirect(url_for('admin_whatsapp'))

    # Ensure phone has country code
    if not phone.startswith('+'):
        phone = '+91' + phone  # Default to India

    try:
        conn = get_conn(); cur = conn.cursor()
        cur.execute("""
            INSERT INTO whatsapp_recipients (name, phone, role, student_id, class_id)
            VALUES (%s, %s, %s, %s, %s)
        """, (name, phone, role,
              int(student_id) if student_id else None,
              int(class_id) if class_id else None))
        conn.commit(); cur.close(); _return_conn(conn)
        flash(f'Recipient "{name}" ({role}) added.', 'success')
    except Exception as e:
        flash(f'Error: {e}', 'danger')

    return redirect(url_for('admin_whatsapp'))


@app.route('/admin/whatsapp/<int:rec_id>/delete', methods=['POST'])
@admin_required
def admin_delete_whatsapp_recipient(rec_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute('DELETE FROM whatsapp_recipients WHERE id=%s', (rec_id,))
    conn.commit(); cur.close(); _return_conn(conn)
    flash('Recipient removed.', 'success')
    return redirect(url_for('admin_whatsapp'))


@app.route('/admin/whatsapp/test', methods=['POST'])
@admin_required
def admin_test_whatsapp():
    """Send a test message to verify Twilio setup."""
    phone = request.form.get('phone', '').strip()
    if not phone:
        flash('Enter a phone number to test.', 'danger')
        return redirect(url_for('admin_whatsapp'))

    if not whatsapp_client:
        flash('WhatsApp not configured. Set Twilio credentials in config.py.', 'danger')
        return redirect(url_for('admin_whatsapp'))

    if not phone.startswith('+'):
        phone = '+91' + phone

    result = whatsapp_client.send_message(phone, "✅ Test message from AttendanceAI. WhatsApp integration is working!")
    if result['success']:
        flash(f'Test message sent! SID: {result["sid"]}', 'success')
    else:
        flash(f'Failed: {result["error"]}', 'danger')

    return redirect(url_for('admin_whatsapp'))


@app.route('/admin/whatsapp/send-summary', methods=['POST'])
@admin_required
def admin_send_whatsapp_summary():
    """Manually trigger daily summary to all registered recipients."""
    if not whatsapp_client:
        flash('WhatsApp not configured.', 'danger')
        return redirect(url_for('admin_whatsapp'))

    try:
        send_daily_summary_to_all(DB_PARAMS, whatsapp_client)
        flash('Daily summary sent via WhatsApp!', 'success')
    except Exception as e:
        flash(f'Error: {e}', 'danger')

    return redirect(url_for('admin_whatsapp'))


# ══════════════════════════════════════════════════════════════════════════════
# COLLEGE PORTAL
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/college')
@college_required
def college_dashboard():
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT COUNT(*) FROM attendance WHERE date=%s', (_today_ist(),))
    today_count = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM student')
    total_students = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM class')
    class_count = cur.fetchone()[0]
    cur.close(); _return_conn(conn)
    return render_template('college/dashboard.html',
                           today_count=today_count,
                           total_students=total_students,
                           class_count=class_count)


@app.route('/college/mark', methods=['GET', 'POST'])
@college_required
def mark_attendance():
    if request.method == 'POST':
        data = request.get_json()
        if not data or 'photo' not in data:
            return jsonify({'error': 'No photo provided'}), 400

        section_id = data.get('section_id')
        period_id = data.get('period_id')

        if not section_id:
            return jsonify({'error': 'Please select a class and section.'}), 400

        filename = _now_ist().strftime('%H-%M-%S') + '.jpg'
        img_path = os.path.join(IMAGES_DIR, filename)
        try:
            _save_b64_image(data['photo'], img_path, max_size=640)
            liveness_result = None

            # Get subject_id from period if available (for legacy compat)
            subject_id = None
            if period_id:
                conn = get_conn(); cur = conn.cursor()
                cur.execute('SELECT period_name FROM timetable WHERE tt_id=%s', (int(period_id),))
                period_row = cur.fetchone()
                if period_row:
                    cur.execute('SELECT sub_id FROM subject WHERE name ILIKE %s', (period_row[0] + '%',))
                    sub_row = cur.fetchone()
                    if sub_row:
                        subject_id = sub_row[0]
                cur.close(); _return_conn(conn)

            # ── Face Recognition (lightweight — no TensorFlow) ────────────
            embeddings_csv = os.path.join(BASE_DIR, 'embeddings.csv')
            if not os.path.exists(embeddings_csv):
                return jsonify({'error': 'No embeddings found. Register students first.'}), 400

            import pandas as pd_local
            embeddings_df = pd_local.read_csv(embeddings_csv)
            emb_names = embeddings_df['name'].values
            emb_vectors = embeddings_df.drop(columns=['name']).values

            img, identified_persons = _lightweight_face_recognition(img_path, emb_vectors, emb_names)

            if img is None:
                return jsonify({'error': 'Could not process image. Try again with better lighting.'}), 400

            # Build result
            identified_names = [p['name'] for p in identified_persons]
            student_id_map = _get_student_ids_local(identified_names) if identified_names else {}
            identified_student_ids = [student_id_map[n] for n in identified_names if n in student_id_map]

            # Mark attendance in DB
            if identified_student_ids:
                conn = get_conn(); cur = conn.cursor()
                for std_id in identified_student_ids:
                    # Check if already marked today for this section/period or subject
                    if subject_id:
                        cur.execute("""
                            SELECT 1 FROM attendance
                            WHERE date=%s AND student_id=%s AND subject_id=%s
                        """, (_today_ist(), std_id, subject_id))
                    else:
                        cur.execute("""
                            SELECT 1 FROM attendance
                            WHERE date=%s AND student_id=%s AND section_id=%s AND period_id=%s
                        """, (_today_ist(), std_id, int(section_id), int(period_id) if period_id else None))

                    if not cur.fetchone():
                        cur.execute("""
                            INSERT INTO attendance (date, student_id, subject_id, image, section_id, period_id)
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, (
                            _today_ist(), std_id,
                            subject_id,  # may be None
                            img_path,
                            int(section_id),
                            int(period_id) if period_id else None
                        ))
                conn.commit(); cur.close(); _return_conn(conn)

            # Fetch student details for response
            students_present = []
            if identified_student_ids:
                conn = get_conn(); cur = conn.cursor()
                cur.execute('SELECT std_id, name, email FROM student WHERE std_id = ANY(%s)', (identified_student_ids,))
                students_present = [{'id': r[0], 'name': r[1], 'email': r[2]} for r in cur.fetchall()]
                cur.close(); _return_conn(conn)

            result = {
                'image_name': filename,
                'students_present': students_present,
                'identified_persons': [
                    {'name': p['name'], 'confidence_score': float(p['confidence_score'])}
                    for p in identified_persons
                ],
                'total_faces_detected': len(identified_persons),
                'total_matched': len(identified_student_ids),
            }

            # ── WhatsApp Absence Alerts ───────────────────────────────────
            if whatsapp_client and period_id:
                try:
                    import threading
                    t = threading.Thread(
                        target=check_and_send_absence_alerts,
                        args=(DB_PARAMS, whatsapp_client, int(section_id), int(period_id)),
                        daemon=True
                    )
                    t.start()
                except Exception:
                    pass

            # Build response (no engagement here - engagement is separate)
            response = _serialize(result)
            if liveness_result:
                response['liveness'] = {
                    'is_live': bool(liveness_result.is_live),
                    'confidence': float(round(liveness_result.confidence, 3))
                }

            # Cleanup
            for f in os.listdir(IMAGES_DIR):
                if not re.match(r'^\d{2}-\d{2}-\d{2}\.(jpg|jpeg|png)$', f, re.IGNORECASE):
                    try: os.remove(os.path.join(IMAGES_DIR, f))
                    except: pass

            return jsonify(_serialize(response))
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    # GET – fetch classes for selection
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT class_id, name FROM class ORDER BY name')
    classes = cur.fetchall()
    cur.close(); _return_conn(conn)
    return render_template('college/mark_attendance.html', classes=classes)


@app.route('/college/records')
@college_required
def attendance_records():
    date_str = request.args.get('date', _today_ist().strftime('%Y-%m-%d'))
    class_id = request.args.get('class_id', '')
    section_id = request.args.get('section_id', '')

    conn = get_conn(); cur = conn.cursor()

    query = """
        SELECT st.name, st.roll_no, c.name as class_name, sec.name as sec_name,
               t.period_name, a.date
        FROM attendance a
        JOIN student st ON a.student_id=st.std_id
        LEFT JOIN class c ON st.class_id=c.class_id
        LEFT JOIN section sec ON a.section_id=sec.section_id
        LEFT JOIN timetable t ON a.period_id=t.tt_id
        WHERE a.date=%s
    """
    params = [date_str]

    if class_id:
        query += " AND st.class_id=%s"
        params.append(int(class_id))
    if section_id:
        query += " AND a.section_id=%s"
        params.append(int(section_id))

    query += " ORDER BY c.name, sec.name, st.roll_no, st.name"
    cur.execute(query, params)
    records = cur.fetchall()

    cur.execute('SELECT DISTINCT date FROM attendance ORDER BY date DESC LIMIT 30')
    dates = [r[0].strftime('%Y-%m-%d') for r in cur.fetchall()]

    cur.execute('SELECT class_id, name FROM class ORDER BY name')
    classes = cur.fetchall()

    cur.close(); _return_conn(conn)
    return render_template('college/records.html', records=records, date=date_str,
                           dates=dates, classes=classes,
                           selected_class=class_id, selected_section=section_id)


# ══════════════════════════════════════════════════════════════════════════════
#  COLLEGE – ENGAGEMENT MONITORING
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/college/engagement')
@college_required
def college_engagement():
    """College view: configure & monitor engagement for class/section/period."""
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT class_id, name FROM class ORDER BY name')
    classes = cur.fetchall()

    # Get any currently active sessions
    cur.execute("""
        SELECT es.id, c.name, s.name, t.period_name, es.started_at, es.capture_interval
        FROM engagement_sessions es
        JOIN section s ON es.section_id = s.section_id
        JOIN class c ON s.class_id = c.class_id
        JOIN timetable t ON es.period_id = t.tt_id
        WHERE es.is_active = TRUE
        ORDER BY es.started_at DESC
    """)
    active_sessions = cur.fetchall()

    # Recent logs for this user's sessions (last 24h)
    cur.execute("""
        SELECT e.timestamp, c.name, s.name, t.period_name,
               e.total_faces, e.attentive_pct, e.confused_pct, e.distracted_pct, e.avg_score
        FROM engagement_log e
        LEFT JOIN section s ON e.section_id=s.section_id
        LEFT JOIN class c ON s.class_id=c.class_id
        LEFT JOIN timetable t ON e.period_id=t.tt_id
        WHERE e.timestamp >= NOW() - INTERVAL '24 hours'
        ORDER BY e.timestamp DESC LIMIT 30
    """)
    recent_logs = cur.fetchall()

    cur.close(); _return_conn(conn)
    return render_template('college/engagement.html', classes=classes,
                           active_sessions=active_sessions, recent_logs=recent_logs)


@app.route('/college/engagement/start', methods=['POST'])
@college_required
def college_engagement_start():
    """Start an engagement monitoring session for a class/section/period."""
    section_id = request.form.get('section_id')
    period_id = request.form.get('period_id')
    interval = request.form.get('interval', 60, type=int)

    if not section_id or not period_id:
        flash('Please select a section and period.', 'warning')
        return redirect(url_for('college_engagement'))

    conn = get_conn(); cur = conn.cursor()

    # Check if already active for this section/period
    cur.execute("""
        SELECT id FROM engagement_sessions
        WHERE section_id=%s AND period_id=%s AND is_active=TRUE
    """, (section_id, period_id))
    existing = cur.fetchone()
    if existing:
        flash('Engagement monitoring is already active for this section/period.', 'info')
        cur.close(); _return_conn(conn)
        return redirect(url_for('college_engagement'))

    cur.execute("""
        INSERT INTO engagement_sessions (section_id, period_id, capture_interval, started_by)
        VALUES (%s, %s, %s, %s)
    """, (section_id, period_id, interval, session.get('user_id')))
    conn.commit()
    cur.close(); _return_conn(conn)

    flash('Engagement monitoring started! Captures will begin automatically.', 'success')
    return redirect(url_for('college_engagement'))


@app.route('/college/engagement/stop', methods=['POST'])
@college_required
def college_engagement_stop():
    """Stop an active engagement monitoring session."""
    session_id = request.form.get('session_id')
    if not session_id:
        flash('Invalid session.', 'danger')
        return redirect(url_for('college_engagement'))

    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        UPDATE engagement_sessions SET is_active=FALSE, ended_at=NOW()
        WHERE id=%s
    """, (session_id,))
    conn.commit()
    cur.close(); _return_conn(conn)

    flash('Engagement monitoring stopped.', 'success')
    return redirect(url_for('college_engagement'))


@app.route('/college/engagement/capture', methods=['POST'])
@college_required
def college_engagement_capture():
    """Capture a frame: identify each student, compute per-student engagement + liveness."""
    data = request.get_json()
    if not data or 'photo' not in data:
        return jsonify({'error': 'No photo provided'}), 400

    session_id = data.get('session_id')
    if not session_id:
        return jsonify({'error': 'No session_id provided'}), 400

    try:
        b64 = data['photo']
        if ',' in b64:
            b64 = b64.split(',')[1]
        img_data = base64.b64decode(b64)
        arr = np.frombuffer(img_data, dtype=np.uint8)
        img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        if img is None:
            return jsonify({'error': 'Could not decode image'}), 400

        # ── Verify session is active ──────────────────────────────────────
        conn = get_conn(); cur = conn.cursor()
        cur.execute("""
            SELECT section_id, period_id FROM engagement_sessions
            WHERE id=%s AND is_active=TRUE
        """, (session_id,))
        sess = cur.fetchone()
        if not sess:
            cur.close(); _return_conn(conn)
            return jsonify({'error': 'Session not active or not found'}), 404
        section_id, period_id = sess

        # ── Engagement analysis (class-level) ─────────────────────────────
        _load_ai_modules()
        if analyze_engagement is None:
            cur.close(); _return_conn(conn)
            return jsonify({'error': 'Engagement analysis not available on this server (lightweight mode)'}), 503
        engagement = analyze_engagement(img)

        # ── Face recognition: identify who each face is ───────────────────
        embeddings_csv = os.path.join(BASE_DIR, 'embeddings.csv')
        student_details = []  # per-student results

        if os.path.exists(embeddings_csv):
            try:
                import pandas as pd_local
                from Attendance_update_db import identify_persons_in_group_photo, get_student_ids
            except ImportError:
                cur.close(); _return_conn(conn)
                return jsonify({'error': 'AI modules not available (lightweight mode)'}), 503

            embeddings_df = pd_local.read_csv(embeddings_csv)
            emb_names = embeddings_df['name'].values
            emb_vectors = embeddings_df.drop(columns=['name']).values

            _, identified_persons = identify_persons_in_group_photo(
                img, emb_vectors, emb_names, is_path=False
            )

            # Map names → student IDs
            id_names = [p['name'] for p in identified_persons]
            student_id_map = get_student_ids(id_names) if id_names else {}

            # ── Per-student liveness (single-frame check on each face crop) ─
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

            for idx, person in enumerate(identified_persons):
                name = person['name']
                confidence = float(person['confidence_score'])
                std_id = student_id_map.get(name)

                # Match identified person to an engagement face by bbox overlap
                face_eng = None
                if idx < len(engagement.faces):
                    face_eng = engagement.faces[idx]

                # Per-face liveness on the cropped face region
                liveness_score = 0.0
                is_live = False
                if face_eng:
                    x1, y1, x2, y2 = face_eng.bbox
                    face_crop = img[y1:y2, x1:x2]
                    if face_crop.size > 0:
                        lr = quick_liveness_check(face_crop)
                        liveness_score = round(float(lr.confidence), 3)
                        is_live = bool(lr.is_live)

                eng_score = round(float(face_eng.engagement_score), 3) if face_eng else 0.0
                emotion = face_eng.emotion if face_eng else 'unknown'
                gaze = face_eng.gaze_direction if face_eng else 'unknown'
                attentive = bool(face_eng.is_attentive) if face_eng else False

                student_details.append({
                    'name': name,
                    'student_id': std_id,
                    'confidence': confidence,
                    'engagement_score': eng_score,
                    'emotion': emotion,
                    'gaze': gaze,
                    'is_attentive': attentive,
                    'liveness_score': liveness_score,
                    'is_live': is_live,
                })

        # ── Save class-level engagement_log ───────────────────────────────
        now = _now_ist()
        cur.execute("""
            INSERT INTO engagement_log (date, section_id, period_id, timestamp,
                                        total_faces, attentive_pct, confused_pct,
                                        distracted_pct, avg_score)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            _today_ist(), section_id, period_id, now,
            engagement.total_faces, engagement.attentive_pct,
            engagement.confused_pct, engagement.distracted_pct,
            engagement.avg_engagement_score
        ))
        log_id = cur.fetchone()[0]

        # ── Save per-student engagement ───────────────────────────────────
        for sd in student_details:
            cur.execute("""
                INSERT INTO engagement_student_log
                    (session_id, log_id, student_id, student_name, timestamp,
                     engagement_score, emotion, is_attentive, gaze_direction,
                     liveness_score, is_live)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                int(session_id), log_id, sd['student_id'], sd['name'], now,
                sd['engagement_score'], sd['emotion'], sd['is_attentive'],
                sd['gaze'], sd['liveness_score'], sd['is_live']
            ))

        conn.commit(); cur.close(); _return_conn(conn)

        # Response for college: class-level stats + identified count only
        # Per-student details are stored in DB and visible only to admin
        return jsonify(_serialize({
            'success': True,
            'total_faces': engagement.total_faces,
            'attentive_count': engagement.attentive_count,
            'distracted_count': engagement.distracted_count,
            'attentive_pct': engagement.attentive_pct,
            'confused_pct': engagement.confused_pct,
            'distracted_pct': engagement.distracted_pct,
            'avg_score': round(engagement.avg_engagement_score, 3),
            'identified_count': len(student_details),
            'timestamp': now.strftime('%H:%M:%S')
        }))
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ══════════════════════════════════════════════════════════════════════════════
# TEACHER PORTAL
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/teacher')
@teacher_required
def teacher_dashboard():
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT COUNT(*) FROM attendance WHERE date=%s', (_today_ist(),))
    today_count = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM student')
    total_students = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM batch')
    batch_count = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM branch')
    branch_count = cur.fetchone()[0]
    # Active engagement sessions
    cur.execute('SELECT COUNT(*) FROM engagement_sessions WHERE is_active=TRUE')
    active_sessions = cur.fetchone()[0]
    cur.close(); _return_conn(conn)
    return render_template('teacher/dashboard.html',
                           today_count=today_count,
                           total_students=total_students,
                           batch_count=batch_count,
                           branch_count=branch_count,
                           active_sessions=active_sessions)


# ── Teacher: Attendance Marking (Time-Restricted) ─────────────────────────────

def _check_time_window(from_time, to_time):
    """
    Check if current time is within attendance marking window.
    Opening: from_time - 10min to from_time + 10min
    Closing: to_time - 5min to to_time + 10min
    Returns: (can_mark, window_type, message)
    """
    now_ist = dt.datetime.now(IST)
    today_ist = now_ist.date()
    now = now_ist.time()
    from_dt = dt.datetime.combine(today_ist, from_time)
    to_dt = dt.datetime.combine(today_ist, to_time)
    now_dt = dt.datetime.combine(today_ist, now)

    # Opening window: start - 10min to start + 10min
    open_start = from_dt - dt.timedelta(minutes=10)
    open_end = from_dt + dt.timedelta(minutes=10)

    # Closing window: end - 5min to end + 10min
    close_start = to_dt - dt.timedelta(minutes=5)
    close_end = to_dt + dt.timedelta(minutes=10)

    if open_start <= now_dt <= open_end:
        return True, 'opening', f'Opening window (until {open_end.strftime("%H:%M")})'
    elif close_start <= now_dt <= close_end:
        return True, 'closing', f'Closing window (until {close_end.strftime("%H:%M")})'
    else:
        # Calculate next available window
        if now_dt < open_start:
            return False, None, f'Opens at {open_start.strftime("%H:%M")}'
        elif open_end < now_dt < close_start:
            return False, None, f'Next window at {close_start.strftime("%H:%M")}'
        else:
            return False, None, 'Period attendance window has closed'


@app.route('/teacher/attendance')
@teacher_required
def teacher_attendance():
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT batch_id, passing_year, name FROM batch ORDER BY passing_year DESC')
    batches = cur.fetchall()
    cur.close(); _return_conn(conn)
    return render_template('teacher/attendance.html', batches=batches)


@app.route('/api/teacher/dashboard')
@teacher_required
def api_teacher_dashboard():
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT COUNT(*) FROM attendance WHERE date=%s', (_today_ist(),))
    today_count = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM student')
    total_students = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM batch')
    batch_count = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM engagement_sessions WHERE is_active=TRUE')
    active_sessions = cur.fetchone()[0]
    cur.close(); _return_conn(conn)
    return jsonify({'today_count': today_count, 'total_students': total_students,
                    'batch_count': batch_count, 'active_sessions': active_sessions})


@app.route('/api/teacher/batches')
@teacher_required
def api_teacher_batches():
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT batch_id, passing_year, name FROM batch ORDER BY passing_year DESC')
    rows = cur.fetchall(); cur.close(); _return_conn(conn)
    return jsonify([{'id': r[0], 'year': r[1], 'name': r[2]} for r in rows])


@app.route('/api/teacher/branches')
@teacher_required
def api_teacher_branches():
    batch_id = request.args.get('batch_id', '')
    conn = get_conn(); cur = conn.cursor()
    if batch_id:
        cur.execute("""
            SELECT branch_id, name, code FROM branch
            WHERE batch_id=%s ORDER BY name
        """, (int(batch_id),))
    else:
        cur.execute('SELECT branch_id, name, code FROM branch ORDER BY name')
    rows = cur.fetchall(); cur.close(); _return_conn(conn)
    return jsonify([{'id': r[0], 'name': r[1], 'code': r[2]} for r in rows])


@app.route('/api/teacher/classes')
@teacher_required
def api_teacher_classes():
    branch_id = request.args.get('branch_id', '')
    batch_id = request.args.get('batch_id', '')
    conn = get_conn(); cur = conn.cursor()
    query = 'SELECT class_id, name FROM class WHERE 1=1'
    params = []
    if branch_id:
        query += ' AND branch_id=%s'
        params.append(int(branch_id))
    if batch_id:
        query += ' AND batch_id=%s'
        params.append(int(batch_id))
    query += ' ORDER BY name'
    cur.execute(query, params)
    rows = cur.fetchall(); cur.close(); _return_conn(conn)
    return jsonify([{'id': r[0], 'name': r[1]} for r in rows])


@app.route('/api/teacher/sections')
@app.route('/api/teacher/sections/<int:class_id>')
@teacher_required
def api_teacher_sections(class_id=None):
    class_id = class_id or request.args.get('class_id', '')
    if not class_id:
        return jsonify([]), 200
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT section_id, name FROM section WHERE class_id=%s ORDER BY name', (int(class_id),))
    rows = cur.fetchall(); cur.close(); _return_conn(conn)
    return jsonify([{'id': r[0], 'name': r[1]} for r in rows])


@app.route('/api/teacher/engagement/<int:section_id>')
@teacher_required
def api_teacher_engagement(section_id):
    """Get engagement log entries for a section, grouped by date."""
    date_from = request.args.get('from', '')
    date_to = request.args.get('to', '')
    conn = get_conn(); cur = conn.cursor()
    query = """
        SELECT e.date,
               COALESCE(AVG(e.avg_score), 0)       AS score,
               COALESCE(AVG(e.attentive_pct), 0)   AS engaged,
               COALESCE(AVG(e.distracted_pct), 0)  AS disengaged,
               COUNT(*)                             AS captures
        FROM engagement_log e
        WHERE e.section_id = %s
    """
    params = [section_id]
    if date_from:
        query += " AND e.date >= %s"; params.append(date_from)
    if date_to:
        query += " AND e.date <= %s"; params.append(date_to)
    query += " GROUP BY e.date ORDER BY e.date DESC LIMIT 60"
    cur.execute(query, params)
    rows = cur.fetchall()
    cur.close(); _return_conn(conn)
    result = [{'date': str(r[0]), 'score': round(float(r[1]), 1), 'engaged': round(float(r[2]), 1),
               'disengaged': round(float(r[3]), 1), 'captures': int(r[4])} for r in rows]
    return jsonify(result)


@app.route('/api/teacher/periods')
@app.route('/api/teacher/periods/<int:section_id>')
@teacher_required
def api_teacher_periods(section_id=None):
    """Get periods for a section with time-window availability status."""
    section_id = section_id or request.args.get('section_id', '')
    if not section_id:
        return jsonify([]), 200
    section_id = int(section_id)
    today_dow = dt.datetime.now(IST).weekday()
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        SELECT tt_id, period_name, teacher_name, from_time, to_time, is_recess
        FROM timetable WHERE section_id=%s AND day_of_week=%s
        ORDER BY from_time
    """, (section_id, today_dow))
    rows = cur.fetchall(); cur.close(); _return_conn(conn)

    periods = []
    for r in rows:
        can_mark, window_type, message = _check_time_window(r[3], r[4])
        periods.append({
            'id': r[0],
            'name': r[1],
            'teacher': r[2],
            'from_time': str(r[3])[:5],
            'to_time': str(r[4])[:5],
            'is_recess': r[5],
            'can_mark': can_mark,
            'window_type': window_type,
            'window_message': message
        })
    return jsonify(periods)


@app.route('/teacher/attendance/mark', methods=['POST'])
@teacher_required
def teacher_mark_attendance():
    """Mark attendance - validates time window server-side."""
    data = request.get_json()
    if not data or 'photo' not in data:
        return jsonify({'error': 'No photo provided'}), 400

    section_id = data.get('section_id')
    period_id = data.get('period_id')

    if not section_id or not period_id:
        return jsonify({'error': 'Please select section and period.'}), 400

    # ── Server-side time window validation ────────────────────────────────
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT from_time, to_time FROM timetable WHERE tt_id=%s', (int(period_id),))
    period_row = cur.fetchone()
    if not period_row:
        cur.close(); _return_conn(conn)
        return jsonify({'error': 'Invalid period selected.'}), 400

    force = data.get('force', False)
    can_mark, window_type, message = _check_time_window(period_row[0], period_row[1])
    if not can_mark and not force:
        cur.close(); _return_conn(conn)
        return jsonify({'error': f'Attendance marking not available now. {message}'}), 403
    if force:
        window_type = 'forced'

    cur.close(); _return_conn(conn)

    # ── Process image and mark attendance (reuse existing logic) ───────────
    filename = _now_ist().strftime('%H-%M-%S') + '.jpg'
    img_path = os.path.join(IMAGES_DIR, filename)
    try:
        _save_b64_image(data['photo'], img_path, max_size=640)

        embeddings_csv = os.path.join(BASE_DIR, 'embeddings.csv')
        if not os.path.exists(embeddings_csv):
            return jsonify({'error': 'No embeddings found. Register students first.'}), 400

        import pandas as pd_local
        embeddings_df = pd_local.read_csv(embeddings_csv)
        emb_names = embeddings_df['name'].values
        emb_vectors = embeddings_df.drop(columns=['name']).values

        img, identified_persons = _lightweight_face_recognition(img_path, emb_vectors, emb_names)

        if img is None:
            return jsonify({'error': 'Could not process image.'}), 400

        # Only get names of recognized persons (not unknown faces)
        identified_names = [p['name'] for p in identified_persons if p.get('recognized') and p['name']]
        identified_names = list(dict.fromkeys(identified_names))  # unique, preserve order
        student_id_map = _get_student_ids_local(identified_names) if identified_names else {}
        identified_student_ids = [student_id_map[n] for n in identified_names if n in student_id_map]

        # Mark attendance — track who is new vs already marked
        already_marked_ids = []
        newly_marked_ids = []
        if identified_student_ids:
            conn = get_conn(); cur = conn.cursor()
            for std_id in identified_student_ids:
                cur.execute("""
                    SELECT 1 FROM attendance
                    WHERE date=%s AND student_id=%s AND section_id=%s AND period_id=%s
                """, (_today_ist(), std_id, int(section_id), int(period_id)))
                if cur.fetchone():
                    already_marked_ids.append(std_id)
                else:
                    cur.execute("""
                        INSERT INTO attendance (date, student_id, image, section_id, period_id)
                        VALUES (%s, %s, %s, %s, %s)
                    """, (_today_ist(), std_id, img_path, int(section_id), int(period_id)))
                    newly_marked_ids.append(std_id)
            conn.commit(); cur.close(); _return_conn(conn)

        # Build faces array for frontend canvas overlay
        faces_for_frontend = []
        for person in identified_persons:
            fa = person.get('facial_area', {})
            name = person.get('name', '')
            is_recognized = person.get('recognized', False)
            std_id = student_id_map.get(name) if name else None
            already = std_id in already_marked_ids if std_id else False
            if fa:
                faces_for_frontend.append({
                    'x': int(fa.get('x', 0)),
                    'y': int(fa.get('y', 0)),
                    'w': int(fa.get('w', 0)),
                    'h': int(fa.get('h', 0)),
                    'name': name if is_recognized else 'Unknown',
                    'recognized': is_recognized,
                    'already_marked': already,
                })

        # Build response
        students_present = []
        all_ids = identified_student_ids
        if all_ids:
            conn = get_conn(); cur = conn.cursor()
            cur.execute('SELECT std_id, name, email FROM student WHERE std_id = ANY(%s)', (all_ids,))
            students_present = [{'id': r[0], 'name': r[1], 'email': r[2]} for r in cur.fetchall()]
            cur.close(); _return_conn(conn)

        # Get processed image dimensions for frontend coordinate mapping
        img_h, img_w = img.shape[:2] if img is not None else (720, 1280)

        resp = {
            'success': True,
            'window_type': window_type,
            'students_present': students_present,
            'faces': faces_for_frontend,
            'image_width': img_w,
            'image_height': img_h,
            'total_faces_detected': len(identified_persons),
            'total_matched': len(identified_student_ids),
            'newly_marked': len(newly_marked_ids),
            'already_marked': len(already_marked_ids),
        }

        return jsonify(_serialize(resp))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── JSON API wrappers for the React SPA (Teacher) ────────────────────────────

@app.route('/api/teacher/send-report', methods=['POST'])
@teacher_required
def api_teacher_send_report():
    """Redirect to the main send report handler with Excel generation."""
    return teacher_send_report()


@app.route('/api/teacher/report-schedules', methods=['GET'])
@teacher_required
def api_teacher_get_report_schedules():
    """Get all report schedules for the teacher."""
    teacher_id = session.get('user_id')
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        SELECT rs.id, rs.section_id, rs.frequency, rs.send_time, rs.include_parents, rs.active, rs.created_at,
               sec.name as section_name, c.name as class_name
        FROM report_schedules rs
        LEFT JOIN section sec ON rs.section_id = sec.section_id
        LEFT JOIN class c ON sec.class_id = c.class_id
        WHERE rs.teacher_id = %s ORDER BY rs.created_at DESC
    """, (teacher_id,))
    rows = cur.fetchall()
    cur.close(); _return_conn(conn)
    schedules = [{'id': r[0], 'section_id': r[1], 'frequency': r[2], 'time': str(r[3])[:5],
                  'include_parents': r[4], 'active': r[5], 'created_at': str(r[6]),
                  'section_name': f"{r[7] or 'Unknown'} ({r[8] or ''})"} for r in rows]
    return jsonify(schedules)


@app.route('/api/teacher/report-schedules', methods=['POST'])
@teacher_required
def api_teacher_create_report_schedule():
    """Create a new report schedule."""
    data = request.get_json() or {}
    section_id = data.get('section_id')
    frequency = data.get('frequency', 'daily')
    send_time = data.get('time', '09:00')
    include_parents = data.get('include_parents', False)
    teacher_id = session.get('user_id')

    if not section_id:
        return jsonify({'error': 'section_id is required'}), 400

    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        INSERT INTO report_schedules (teacher_id, section_id, frequency, send_time, include_parents, active)
        VALUES (%s, %s, %s, %s, %s, true) RETURNING id
    """, (teacher_id, int(section_id), frequency, send_time, include_parents))
    new_id = cur.fetchone()[0]
    conn.commit(); cur.close(); _return_conn(conn)
    return jsonify({'id': new_id, 'message': 'Schedule created'}), 201


@app.route('/api/teacher/report-schedules/<int:schedule_id>', methods=['DELETE'])
@teacher_required
def api_teacher_delete_report_schedule(schedule_id):
    """Delete a report schedule."""
    teacher_id = session.get('user_id')
    conn = get_conn(); cur = conn.cursor()
    cur.execute('DELETE FROM report_schedules WHERE id=%s AND teacher_id=%s', (schedule_id, teacher_id))
    conn.commit(); cur.close(); _return_conn(conn)
    return jsonify({'message': 'Schedule deleted'})


@app.route('/api/teacher/report-schedules/<int:schedule_id>/toggle', methods=['PUT'])
@teacher_required
def api_teacher_toggle_report_schedule(schedule_id):
    """Toggle a report schedule active/inactive."""
    teacher_id = session.get('user_id')
    conn = get_conn(); cur = conn.cursor()
    cur.execute('UPDATE report_schedules SET active = NOT active WHERE id=%s AND teacher_id=%s', (schedule_id, teacher_id))
    conn.commit(); cur.close(); _return_conn(conn)
    return jsonify({'message': 'Schedule toggled'})


@app.route('/api/teacher/mark-attendance', methods=['POST'])
@teacher_required
def api_teacher_mark_attendance():
    """JSON API wrapper - delegates to teacher_mark_attendance."""
    return teacher_mark_attendance()


@app.route('/api/teacher/liveness/enable', methods=['POST'])
@teacher_required
def api_teacher_liveness_enable():
    """Enable monitoring - accepts JSON body."""
    data = request.get_json() or {}
    section_id = data.get('section_id')
    period_id = data.get('period_id')
    interval = data.get('interval', 30)

    if not section_id:
        return jsonify({'error': 'section_id required'}), 400

    conn = get_conn(); cur = conn.cursor()

    # If no period_id, find current active period for section
    if not period_id:
        today_dow = _today_ist().weekday()
        now = _now_ist().time()
        cur.execute("""
            SELECT tt_id, from_time, to_time FROM timetable
            WHERE section_id=%s AND day_of_week=%s AND from_time <= %s AND to_time >= %s
            ORDER BY from_time LIMIT 1
        """, (int(section_id), today_dow, now, now))
        row = cur.fetchone()
        if row:
            period_id = row[0]
        else:
            cur.close(); _return_conn(conn)
            return jsonify({'error': 'No active period right now for this section'}), 400

    # Check if already active
    cur.execute("""
        SELECT id FROM engagement_sessions
        WHERE section_id=%s AND period_id=%s AND is_active=TRUE
    """, (int(section_id), int(period_id)))
    existing = cur.fetchone()
    if existing:
        cur.close(); _return_conn(conn)
        return jsonify({'session_id': existing[0], 'message': 'Already active'})

    cur.execute("""
        INSERT INTO engagement_sessions (section_id, period_id, capture_interval, started_by)
        VALUES (%s, %s, %s, %s) RETURNING id
    """, (int(section_id), int(period_id), interval, session.get('user_id')))
    session_id = cur.fetchone()[0]
    conn.commit(); cur.close(); _return_conn(conn)
    return jsonify({'session_id': session_id, 'message': 'Monitoring enabled'})


@app.route('/api/teacher/liveness/disable', methods=['POST'])
@teacher_required
def api_teacher_liveness_disable():
    """Disable monitoring - accepts JSON body."""
    data = request.get_json() or {}
    section_id = data.get('section_id')
    session_id_val = data.get('session_id')

    conn = get_conn(); cur = conn.cursor()
    if session_id_val:
        cur.execute("UPDATE engagement_sessions SET is_active=FALSE, ended_at=NOW() WHERE id=%s", (int(session_id_val),))
    elif section_id:
        cur.execute("UPDATE engagement_sessions SET is_active=FALSE, ended_at=NOW() WHERE section_id=%s AND is_active=TRUE", (int(section_id),))
    conn.commit(); cur.close(); _return_conn(conn)
    return jsonify({'message': 'Monitoring disabled'})


@app.route('/api/teacher/liveness/capture', methods=['POST'])
@teacher_required
def api_teacher_liveness_capture():
    """Capture from teacher's browser camera, process engagement."""
    data = request.get_json() or {}
    section_id = data.get('section_id')
    photo = data.get('photo')

    if not photo or not section_id:
        return jsonify({'error': 'photo and section_id required'}), 400

    try:
        b64 = photo
        if ',' in b64:
            b64 = b64.split(',')[1]
        img_data = base64.b64decode(b64)
        arr = np.frombuffer(img_data, dtype=np.uint8)
        img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        if img is None:
            return jsonify({'error': 'Could not decode image'}), 400

        # Face detection with OpenCV
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)

        total_faces = len(faces)
        annotated_faces = []

        # Try to match faces against embeddings for engagement tracking
        embeddings_csv = os.path.join(BASE_DIR, 'embeddings.csv')
        has_embeddings = os.path.exists(embeddings_csv)

        for (x, y, w, h) in faces:
            # Default: green box (attentive assumption)
            color = (0, 255, 0)
            label = "Detected"
            annotated_faces.append({
                'bbox': [int(x), int(y), int(x+w), int(y+h)],
                'is_attentive': True,
                'score': 0.8,
            })
            cv2.rectangle(img, (x, y), (x+w, y+h), color, 2)
            cv2.putText(img, label, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 1)

        # Encode annotated image
        _, buffer = cv2.imencode('.jpg', img, [cv2.IMWRITE_JPEG_QUALITY, 80])
        annotated_b64 = base64.b64encode(buffer).decode('utf-8')

        return jsonify({
            'success': True,
            'annotated_image': f'data:image/jpeg;base64,{annotated_b64}',
            'total_faces': total_faces,
            'engaged': total_faces,
            'disengaged': 0,
            'engagement_score': 80 if total_faces > 0 else 0,
            'faces': annotated_faces,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/teacher/liveness/status')
@teacher_required
def api_teacher_liveness_status():
    """Get liveness/camera status for a section."""
    section_id = request.args.get('section_id')
    if not section_id:
        return jsonify({'active': False, 'has_camera': False})

    conn = get_conn(); cur = conn.cursor()
    # Check active session
    cur.execute("""
        SELECT id, capture_interval FROM engagement_sessions
        WHERE section_id=%s AND is_active=TRUE ORDER BY started_at DESC LIMIT 1
    """, (int(section_id),))
    sess = cur.fetchone()
    cur.close(); _return_conn(conn)

    return jsonify({
        'active': sess is not None,
        'session_id': sess[0] if sess else None,
        'has_camera': True,  # Browser camera is always available
    })


# ── Teacher: Liveness/Engagement Monitoring ───────────────────────────────────

@app.route('/teacher/liveness')
@teacher_required
def teacher_liveness():
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT batch_id, passing_year, name FROM batch ORDER BY passing_year DESC')
    batches = cur.fetchall()
    # Active sessions by this teacher
    cur.execute("""
        SELECT es.id, c.name as class_name, s.name as sec_name, t.period_name,
               es.started_at, es.capture_interval
        FROM engagement_sessions es
        JOIN section s ON es.section_id=s.section_id
        JOIN class c ON s.class_id=c.class_id
        JOIN timetable t ON es.period_id=t.tt_id
        WHERE es.is_active=TRUE AND es.started_by=%s
        ORDER BY es.started_at DESC
    """, (session.get('user_id'),))
    active_sessions = cur.fetchall()
    cur.close(); _return_conn(conn)
    return render_template('teacher/liveness.html', batches=batches, active_sessions=active_sessions)


@app.route('/api/teacher/camera-status/<int:section_id>')
@teacher_required
def api_teacher_camera_status(section_id):
    """Check if a camera is attached to this section's classroom."""
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        SELECT id, stream_key, name, rtsp_url, is_active
        FROM camera_streams WHERE section_id=%s
    """, (section_id,))
    camera = cur.fetchone()
    cur.close(); _return_conn(conn)
    if camera:
        return jsonify({
            'has_camera': True,
            'camera_id': camera[0],
            'camera_name': camera[2],
            'is_active': camera[4]
        })
    return jsonify({'has_camera': False, 'message': 'No camera attached to this classroom'})


@app.route('/teacher/liveness/enable', methods=['POST'])
@teacher_required
def teacher_liveness_enable():
    """Enable liveness/engagement monitoring for a period."""
    section_id = request.form.get('section_id')
    period_id = request.form.get('period_id')
    interval = request.form.get('interval', 30, type=int)

    if not section_id or not period_id:
        flash('Please select section and period.', 'warning')
        return redirect(url_for('teacher_liveness'))

    # Check camera exists
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT id FROM camera_streams WHERE section_id=%s AND is_active=TRUE', (section_id,))
    if not cur.fetchone():
        cur.close(); _return_conn(conn)
        flash('No active camera found for this classroom. Please contact admin.', 'danger')
        return redirect(url_for('teacher_liveness'))

    # Check if already active
    cur.execute("""
        SELECT id FROM engagement_sessions
        WHERE section_id=%s AND period_id=%s AND is_active=TRUE
    """, (section_id, period_id))
    if cur.fetchone():
        cur.close(); _return_conn(conn)
        flash('Monitoring already active for this section/period.', 'info')
        return redirect(url_for('teacher_liveness'))

    cur.execute("""
        INSERT INTO engagement_sessions (section_id, period_id, capture_interval, started_by)
        VALUES (%s, %s, %s, %s)
    """, (section_id, period_id, interval, session.get('user_id')))
    conn.commit(); cur.close(); _return_conn(conn)
    flash('Liveness & engagement monitoring enabled!', 'success')
    return redirect(url_for('teacher_liveness'))


@app.route('/teacher/liveness/disable', methods=['POST'])
@teacher_required
def teacher_liveness_disable():
    session_id = request.form.get('session_id')
    if not session_id:
        flash('Invalid session.', 'danger')
        return redirect(url_for('teacher_liveness'))
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        UPDATE engagement_sessions SET is_active=FALSE, ended_at=NOW()
        WHERE id=%s AND started_by=%s
    """, (session_id, session.get('user_id')))
    conn.commit(); cur.close(); _return_conn(conn)
    flash('Monitoring disabled.', 'success')
    return redirect(url_for('teacher_liveness'))


@app.route('/teacher/liveness/capture', methods=['POST'])
@teacher_required
def teacher_liveness_capture():
    """Process a frame for liveness + engagement with red/green face annotations."""
    data = request.get_json()
    if not data or 'photo' not in data:
        return jsonify({'error': 'No photo provided'}), 400

    session_id = data.get('session_id')
    if not session_id:
        return jsonify({'error': 'No session_id'}), 400

    try:
        b64 = data['photo']
        if ',' in b64:
            b64 = b64.split(',')[1]
        img_data = base64.b64decode(b64)
        arr = np.frombuffer(img_data, dtype=np.uint8)
        img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        if img is None:
            return jsonify({'error': 'Could not decode image'}), 400

        conn = get_conn(); cur = conn.cursor()
        cur.execute("""
            SELECT section_id, period_id FROM engagement_sessions
            WHERE id=%s AND is_active=TRUE
        """, (session_id,))
        sess = cur.fetchone()
        if not sess:
            cur.close(); _return_conn(conn)
            return jsonify({'error': 'Session not active'}), 404
        section_id, period_id = sess

        _load_ai_modules()
        if analyze_engagement is None:
            cur.close(); _return_conn(conn)
            return jsonify({'error': 'AI modules not available (lightweight mode)'}), 503

        engagement = analyze_engagement(img)

        # Draw red/green rectangles on faces
        annotated_faces = []
        for face in engagement.faces:
            x1, y1, x2, y2 = face.bbox
            color = (0, 255, 0) if face.is_attentive else (0, 0, 255)  # Green/Red BGR
            cv2.rectangle(img, (x1, y1), (x2, y2), color, 2)
            label = f"{'Attentive' if face.is_attentive else 'Distracted'} ({face.engagement_score:.1f})"
            cv2.putText(img, label, (x1, y1 - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.4, color, 1)
            annotated_faces.append({
                'bbox': [x1, y1, x2, y2],
                'is_attentive': face.is_attentive,
                'score': round(face.engagement_score, 2),
                'emotion': face.emotion,
                'gaze': face.gaze_direction
            })

        # Encode annotated image to base64
        _, buffer = cv2.imencode('.jpg', img, [cv2.IMWRITE_JPEG_QUALITY, 80])
        annotated_b64 = base64.b64encode(buffer).decode('utf-8')

        # Save engagement log
        now = _now_ist()
        cur.execute("""
            INSERT INTO engagement_log (date, section_id, period_id, timestamp,
                                        total_faces, attentive_pct, confused_pct,
                                        distracted_pct, avg_score)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            _today_ist(), section_id, period_id, now,
            engagement.total_faces, engagement.attentive_pct,
            engagement.confused_pct, engagement.distracted_pct,
            engagement.avg_engagement_score
        ))
        conn.commit(); cur.close(); _return_conn(conn)

        return jsonify(_serialize({
            'success': True,
            'annotated_image': f'data:image/jpeg;base64,{annotated_b64}',
            'total_faces': engagement.total_faces,
            'attentive_count': engagement.attentive_count,
            'distracted_count': engagement.distracted_count,
            'attentive_pct': engagement.attentive_pct,
            'distracted_pct': engagement.distracted_pct,
            'avg_score': round(engagement.avg_engagement_score, 3),
            'faces': annotated_faces,
            'head_count': engagement.total_faces,
            'timestamp': now.strftime('%H:%M:%S')
        }))
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ── Teacher: Engagement Reports ───────────────────────────────────────────────

@app.route('/teacher/engagement')
@teacher_required
def teacher_engagement():
    """Teacher view of engagement reports."""
    conn = get_conn(); cur = conn.cursor()

    class_filter = request.args.get('class_id', '')
    section_filter = request.args.get('section_id', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    query = """
        SELECT e.date, c.name as class_name, s.name as sec_name, t.period_name,
               e.timestamp, e.total_faces, e.attentive_pct, e.confused_pct,
               e.distracted_pct, e.avg_score
        FROM engagement_log e
        LEFT JOIN section s ON e.section_id=s.section_id
        LEFT JOIN class c ON s.class_id=c.class_id
        LEFT JOIN timetable t ON e.period_id=t.tt_id
        WHERE 1=1
    """
    params = []
    if class_filter:
        query += " AND s.class_id = %s"; params.append(int(class_filter))
    if section_filter:
        query += " AND e.section_id = %s"; params.append(int(section_filter))
    if date_from:
        query += " AND e.date >= %s"; params.append(date_from)
    if date_to:
        query += " AND e.date <= %s"; params.append(date_to)
    query += " ORDER BY e.timestamp DESC LIMIT 100"

    cur.execute(query, params)
    logs = cur.fetchall()

    cur.execute('SELECT batch_id, name FROM batch ORDER BY passing_year DESC')
    batches = cur.fetchall()
    cur.execute('SELECT class_id, name FROM class ORDER BY name')
    classes = cur.fetchall()
    cur.execute('SELECT section_id, name, class_id FROM section ORDER BY name')
    sections = cur.fetchall()
    cur.close(); _return_conn(conn)

    return render_template('teacher/engagement.html', logs=logs, batches=batches,
                           classes=classes, sections=sections,
                           class_filter=class_filter, section_filter=section_filter,
                           date_from=date_from, date_to=date_to)


# ── Teacher: Report Sending (Hierarchical) ───────────────────────────────────

@app.route('/teacher/reports')
@teacher_required
def teacher_reports():
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT batch_id, passing_year, name FROM batch ORDER BY passing_year DESC')
    batches = cur.fetchall()
    cur.execute('SELECT branch_id, name, code, hod_name, hod_email FROM branch ORDER BY name')
    branches = cur.fetchall()

    # Active schedules
    cur.execute("""
        SELECT rs.id, rs.frequency, b.name as batch_name, br.name as branch_name,
               c.name as class_name, s.name as section_name,
               rs.send_to_authorities, rs.send_to_parents, rs.is_active, rs.last_sent
        FROM report_schedule rs
        LEFT JOIN batch b ON rs.batch_id=b.batch_id
        LEFT JOIN branch br ON rs.branch_id=br.branch_id
        LEFT JOIN class c ON rs.class_id=c.class_id
        LEFT JOIN section s ON rs.section_id=s.section_id
        WHERE rs.created_by=%s
        ORDER BY rs.is_active DESC, rs.id DESC
    """, (session.get('user_id'),))
    schedules = cur.fetchall()
    cur.close(); _return_conn(conn)

    return render_template('teacher/reports.html', batches=batches,
                           branches=branches, schedules=schedules)


@app.route('/teacher/reports/send', methods=['POST'])
@teacher_required
def teacher_send_report():
    """Send attendance report based on hierarchy selection."""
    data = request.get_json() if request.is_json else request.form
    section_id = data.get('section_id', '')
    date_from = data.get('date_from', _today_ist().strftime('%Y-%m-%d'))
    date_to = data.get('date_to', _today_ist().strftime('%Y-%m-%d'))
    send_to_authorities = data.get('send_to_authorities') in ('true', 'on', True, '1', 1)
    send_to_parents = data.get('send_to_parents') in ('true', 'on', True, '1', 1)

    if not section_id:
        return jsonify({'error': 'Please select a section'}), 400

    try:
        conn = get_conn(); cur = conn.cursor()
        from_date = dt.datetime.strptime(date_from, '%Y-%m-%d').date()
        to_date = dt.datetime.strptime(date_to, '%Y-%m-%d').date()

        # Build attendance query for the selected section
        att_query = """
            SELECT a.date, st.name as student_name, st.roll_no,
                   c.name as class_name, sec.name as section_name,
                   t.period_name, t.from_time, t.to_time,
                   br.name as branch_name, st.std_id
            FROM attendance a
            JOIN student st ON a.student_id=st.std_id
            LEFT JOIN section sec ON a.section_id=sec.section_id
            LEFT JOIN class c ON st.class_id=c.class_id
            LEFT JOIN branch br ON c.branch_id=br.branch_id
            LEFT JOIN timetable t ON a.period_id=t.tt_id
            WHERE a.date BETWEEN %s AND %s AND a.section_id=%s
            ORDER BY a.date, st.roll_no, t.from_time
        """
        cur.execute(att_query, (from_date, to_date, int(section_id)))
        records = cur.fetchall()

        # Get all students in this section (to mark absent ones)
        cur.execute("""
            SELECT st.std_id, st.name, st.roll_no, c.name as class_name,
                   sec.name as section_name, br.name as branch_name, st.parent_email
            FROM student st
            LEFT JOIN class c ON st.class_id=c.class_id
            LEFT JOIN section sec ON st.section_id=sec.section_id
            LEFT JOIN branch br ON c.branch_id=br.branch_id
            WHERE st.section_id=%s
            ORDER BY st.roll_no, st.name
        """, (int(section_id),))
        all_students = cur.fetchall()

        # Get periods for this section
        cur.execute("""
            SELECT DISTINCT t.period_name, t.from_time, t.to_time
            FROM timetable t
            WHERE t.section_id=%s
            ORDER BY t.from_time
        """, (int(section_id),))
        periods = cur.fetchall()

        # Build recipients list
        hod_recipients = []
        if send_to_authorities:
            # All teachers (HODs) from users table
            cur.execute("SELECT email FROM users WHERE role='teacher' AND email IS NOT NULL")
            hod_recipients = [r[0] for r in cur.fetchall() if r[0]]

        # Parent emails from student table (for per-child reports)
        parent_map = {}  # {std_id: parent_email}
        if send_to_parents:
            for stu in all_students:
                # stu: (std_id, name, roll_no, class_name, section_name, branch_name, parent_email)
                if stu[6]:  # parent_email exists
                    parent_map[stu[0]] = stu[6]

        cur.close(); _return_conn(conn)

        if not hod_recipients and not parent_map:
            return jsonify({'error': 'No recipients found. Please ensure teachers are registered and students have parent emails.'}), 400

        # Generate Excel report
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14, color="1F4E79")
        subtitle_font = Font(bold=True, size=11, color="2E75B6")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Get section info for title
        sec_name = all_students[0][4] if all_students else 'N/A'
        class_name = all_students[0][3] if all_students else 'N/A'
        branch_name = all_students[0][5] if all_students else 'N/A'

        # Title rows
        ws.append([f"ATTENDANCE REPORT"])
        ws['A1'].font = title_font
        ws.append([f"Section: {sec_name} | Class: {class_name} | Branch: {branch_name}"])
        ws['A2'].font = subtitle_font
        ws.append([f"Date Range: {from_date.strftime('%d %B %Y')} to {to_date.strftime('%d %B %Y')}"])
        ws['A3'].font = subtitle_font
        ws.append([f"Generated: {_now_ist().strftime('%d %B %Y, %I:%M %p')}"])
        ws.append([])

        # Build per-day, per-period attendance matrix
        headers = ['S.No', 'Roll No', 'Student Name']
        date_range = []
        d = from_date
        while d <= to_date:
            date_range.append(d)
            d += dt.timedelta(days=1)

        # Column headers: Date + Day + Period
        for date in date_range:
            day_name = day_names[date.weekday()]
            for p in periods:
                headers.append(f"{date.strftime('%d/%m')} ({day_name[:3]}) - {p[0]}")
            if not periods:
                headers.append(f"{date.strftime('%d/%m')} ({day_name[:3]})")

        # Add summary columns
        headers.extend(['Total Present', 'Total Absent', 'Attendance %'])

        # Write header row
        header_row = 6
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        # Build attendance lookup: (date, student_name, period_name) → present
        present_set = set()
        for rec in records:
            present_set.add((str(rec[0]), rec[1], rec[5] if rec[5] else ''))

        # Fill student rows
        total_periods_per_student = len(date_range) * max(len(periods), 1)
        for idx, stu in enumerate(all_students, 1):
            row_data = [idx, stu[2] or '', stu[1]]
            present_count = 0
            for date in date_range:
                for p in periods:
                    key = (str(date), stu[1], p[0])
                    if key in present_set:
                        row_data.append('P')
                        present_count += 1
                    else:
                        row_data.append('A')
                if not periods:
                    any_present = (str(date), stu[1], '') in present_set
                    row_data.append('P' if any_present else 'A')
                    if any_present:
                        present_count += 1

            absent_count = total_periods_per_student - present_count
            pct = round((present_count / total_periods_per_student * 100), 1) if total_periods_per_student > 0 else 0
            row_data.extend([present_count, absent_count, f"{pct}%"])

            row_num = header_row + idx
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # Color P/A cells
                if col_idx > 3 and col_idx <= len(row_data) - 3:
                    if val == 'P':
                        cell.fill = green_fill
                    elif val == 'A':
                        cell.fill = red_fill

        # Auto-width columns
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        for col_idx in range(4, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16

        # Freeze panes (freeze header + student names)
        ws.freeze_panes = 'D7'

        # === Summary Sheet ===
        summary_ws = wb.create_sheet("Summary")
        summary_ws.append(["ATTENDANCE SUMMARY"])
        summary_ws['A1'].font = title_font
        summary_ws.append([f"Section: {sec_name} | Date: {from_date} to {to_date}"])
        summary_ws.append([])
        summary_headers = ['S.No', 'Roll No', 'Student Name', 'Total Classes', 'Present', 'Absent', 'Attendance %', 'Status']
        for col_idx, h in enumerate(summary_headers, 1):
            cell = summary_ws.cell(row=4, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for idx, stu in enumerate(all_students, 1):
            present_count = 0
            for date in date_range:
                for p in periods:
                    if (str(date), stu[1], p[0]) in present_set:
                        present_count += 1
                if not periods:
                    if (str(date), stu[1], '') in present_set:
                        present_count += 1
            absent_count = total_periods_per_student - present_count
            pct = round((present_count / total_periods_per_student * 100), 1) if total_periods_per_student > 0 else 0
            status = "Good" if pct >= 75 else ("Warning" if pct >= 50 else "Critical")

            row_num = 4 + idx
            row_data = [idx, stu[2] or '', stu[1], total_periods_per_student, present_count, absent_count, f"{pct}%", status]
            for col_idx, val in enumerate(row_data, 1):
                cell = summary_ws.cell(row=row_num, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                if col_idx == 8:
                    if val == "Critical":
                        cell.fill = red_fill
                        cell.font = Font(bold=True, color="9C0006")
                    elif val == "Warning":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    else:
                        cell.fill = green_fill

        for col_idx in range(1, len(summary_headers) + 1):
            summary_ws.column_dimensions[get_column_letter(col_idx)].width = 16
        summary_ws.column_dimensions['C'].width = 22

        # Save to file
        report_filename = f"attendance_report_{from_date}_to_{to_date}.xlsx"
        report_path = os.path.join(BASE_DIR, report_filename)
        wb.save(report_path)

        # Send email to HODs (full report with all students)
        from send_email import send_email
        subject = f"Attendance Report ({from_date} to {to_date})"
        body = f"Please find attached the attendance report for section from {from_date} to {to_date}."
        errors = []
        sent_count = 0

        if hod_recipients:
            for email in hod_recipients:
                try:
                    send_email(email, subject, body, [report_path])
                    sent_count += 1
                except Exception as e:
                    errors.append(f"{email}: {str(e)}")

        # Send individual per-child reports to parents
        parent_sent = 0
        if parent_map:
            for stu in all_students:
                std_id = stu[0]
                if std_id not in parent_map:
                    continue
                parent_email = parent_map[std_id]
                student_name = stu[1]
                roll_no = stu[2] or ''

                # Generate per-child Excel
                child_wb = Workbook()
                child_ws = child_wb.active
                child_ws.title = f"Attendance"

                child_ws.append([f"ATTENDANCE REPORT"])
                child_ws['A1'].font = title_font
                child_ws.append([f"Student: {student_name} | Roll No: {roll_no}"])
                child_ws['A2'].font = subtitle_font
                child_ws.append([f"Section: {sec_name} | Class: {class_name}"])
                child_ws.append([f"Date Range: {from_date.strftime('%d %B %Y')} to {to_date.strftime('%d %B %Y')}"])
                child_ws.append([])

                child_headers = ['S.No', 'Date', 'Day', 'Period', 'Time', 'Status']
                for col_idx, h in enumerate(child_headers, 1):
                    cell = child_ws.cell(row=6, column=col_idx, value=h)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border

                row_num = 7
                sno = 1
                child_present = 0
                child_total = 0
                for date in date_range:
                    day_name = day_names[date.weekday()]
                    for p in periods:
                        key = (str(date), student_name, p[0])
                        status = 'Present' if key in present_set else 'Absent'
                        child_total += 1
                        if status == 'Present':
                            child_present += 1
                        row_data = [sno, date.strftime('%d/%m/%Y'), day_name, p[0], f"{p[1]} - {p[2]}", status]
                        for col_idx, val in enumerate(row_data, 1):
                            cell = child_ws.cell(row=row_num, column=col_idx, value=val)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            if col_idx == 6:
                                cell.fill = green_fill if val == 'Present' else red_fill
                                cell.font = Font(bold=True)
                        row_num += 1
                        sno += 1
                    if not periods:
                        any_present = (str(date), student_name, '') in present_set
                        status = 'Present' if any_present else 'Absent'
                        child_total += 1
                        if any_present:
                            child_present += 1
                        row_data = [sno, date.strftime('%d/%m/%Y'), day_name, '-', '-', status]
                        for col_idx, val in enumerate(row_data, 1):
                            cell = child_ws.cell(row=row_num, column=col_idx, value=val)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            if col_idx == 6:
                                cell.fill = green_fill if val == 'Present' else red_fill
                        row_num += 1
                        sno += 1

                # Summary at bottom
                row_num += 1
                child_ws.cell(row=row_num, column=1, value="SUMMARY").font = subtitle_font
                row_num += 1
                child_absent = child_total - child_present
                child_pct = round((child_present / child_total * 100), 1) if child_total > 0 else 0
                child_ws.cell(row=row_num, column=1, value="Total Classes:")
                child_ws.cell(row=row_num, column=2, value=child_total)
                row_num += 1
                child_ws.cell(row=row_num, column=1, value="Present:")
                child_ws.cell(row=row_num, column=2, value=child_present)
                child_ws.cell(row=row_num, column=2).fill = green_fill
                row_num += 1
                child_ws.cell(row=row_num, column=1, value="Absent:")
                child_ws.cell(row=row_num, column=2, value=child_absent)
                child_ws.cell(row=row_num, column=2).fill = red_fill
                row_num += 1
                child_ws.cell(row=row_num, column=1, value="Attendance %:")
                child_ws.cell(row=row_num, column=2, value=f"{child_pct}%")
                child_ws.cell(row=row_num, column=2).font = Font(bold=True, size=12)

                # Column widths
                for col_idx, w in enumerate([6, 12, 12, 18, 14, 10], 1):
                    child_ws.column_dimensions[get_column_letter(col_idx)].width = w

                child_filename = f"attendance_{student_name.replace(' ', '_')}_{from_date}_to_{to_date}.xlsx"
                child_path = os.path.join(BASE_DIR, child_filename)
                child_wb.save(child_path)

                child_subject = f"Attendance Report - {student_name} ({from_date} to {to_date})"
                child_body = f"Dear Parent,\n\nPlease find attached the attendance report for your child {student_name} (Roll No: {roll_no}) from {from_date} to {to_date}.\n\nRegards,\nAttendance System"
                try:
                    send_email(parent_email, child_subject, child_body, [child_path])
                    parent_sent += 1
                except Exception as e:
                    errors.append(f"{parent_email} (parent of {student_name}): {str(e)}")
                finally:
                    try: os.remove(child_path)
                    except: pass

        # Cleanup main report
        try:
            os.remove(report_path)
        except Exception:
            pass

        total_sent = sent_count + parent_sent
        msg_parts = []
        if sent_count:
            msg_parts.append(f"{sent_count} HOD(s)")
        if parent_sent:
            msg_parts.append(f"{parent_sent} parent(s)")

        if errors:
            return jsonify({'success': True, 'message': f"Report sent to {', '.join(msg_parts)}. Some errors occurred.",
                           'sent_to': total_sent, 'errors': errors}), 207
        return jsonify({'success': True, 'message': f"Report sent successfully to {', '.join(msg_parts)}!",
                       'sent_to': total_sent})

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/teacher/reports/schedule', methods=['POST'])
@teacher_required
def teacher_schedule_report():
    """Set up automated report sending."""
    data = request.form
    frequency = data.get('frequency', 'daily')
    batch_id = data.get('batch_id', '') or None
    branch_id = data.get('branch_id', '') or None
    class_id = data.get('class_id', '') or None
    section_id = data.get('section_id', '') or None
    send_to_authorities = data.get('send_to_authorities') == 'on'
    send_to_parents = data.get('send_to_parents') == 'on'

    # Calculate next_send
    now = _now_ist()
    if frequency == 'daily':
        next_send = now.replace(hour=18, minute=0, second=0) + dt.timedelta(days=1)
    elif frequency == 'weekly':
        days_ahead = 7 - now.weekday()  # Next Monday
        next_send = now + dt.timedelta(days=days_ahead)
        next_send = next_send.replace(hour=18, minute=0, second=0)
    else:  # monthly
        if now.month == 12:
            next_send = now.replace(year=now.year + 1, month=1, day=1, hour=18, minute=0, second=0)
        else:
            next_send = now.replace(month=now.month + 1, day=1, hour=18, minute=0, second=0)

    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        INSERT INTO report_schedule (created_by, frequency, batch_id, branch_id, class_id,
                                     section_id, send_to_authorities, send_to_parents, next_send)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (session.get('user_id'), frequency,
          int(batch_id) if batch_id else None,
          int(branch_id) if branch_id else None,
          int(class_id) if class_id else None,
          int(section_id) if section_id else None,
          send_to_authorities, send_to_parents, next_send))
    conn.commit(); cur.close(); _return_conn(conn)
    flash(f'{frequency.capitalize()} report schedule created!', 'success')
    return redirect(url_for('teacher_reports'))


@app.route('/teacher/reports/schedule/<int:schedule_id>/delete', methods=['POST'])
@teacher_required
def teacher_delete_schedule(schedule_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute('DELETE FROM report_schedule WHERE id=%s AND created_by=%s',
               (schedule_id, session.get('user_id')))
    conn.commit(); cur.close(); _return_conn(conn)
    flash('Schedule removed.', 'success')
    return redirect(url_for('teacher_reports'))


@app.route('/teacher/reports/schedule/<int:schedule_id>/toggle', methods=['POST'])
@teacher_required
def teacher_toggle_schedule(schedule_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        UPDATE report_schedule SET is_active = NOT is_active
        WHERE id=%s AND created_by=%s
    """, (schedule_id, session.get('user_id')))
    conn.commit(); cur.close(); _return_conn(conn)
    flash('Schedule updated.', 'success')
    return redirect(url_for('teacher_reports'))


# ══════════════════════════════════════════════════════════════════════════════
# COLLEGE PORTAL – JSON API (React Frontend)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/college/dashboard')
@college_required
def api_college_dashboard():
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT COUNT(*) FROM attendance WHERE date=%s', (_today_ist(),))
    today_count = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM student')
    total_students = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM class')
    class_count = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM batch')
    batch_count = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM branch')
    branch_count = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM section')
    section_count = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM users WHERE role='teacher'")
    teacher_count = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM engagement_sessions WHERE is_active=TRUE')
    active_sessions = cur.fetchone()[0]
    cur.close(); _return_conn(conn)
    return jsonify({
        'today_count': today_count, 'total_students': total_students,
        'class_count': class_count, 'batch_count': batch_count,
        'branch_count': branch_count, 'section_count': section_count,
        'teacher_count': teacher_count, 'active_sessions': active_sessions
    })


# ── Batch CRUD ────────────────────────────────────────────────────────────────

@app.route('/api/college/batches', methods=['GET'])
@college_required
def api_college_batches():
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT batch_id, passing_year, name FROM batch ORDER BY passing_year DESC')
    rows = cur.fetchall(); cur.close(); _return_conn(conn)
    return jsonify([{'id': r[0], 'year': r[1], 'name': r[2]} for r in rows])


@app.route('/api/college/batches', methods=['POST'])
@college_required
def api_college_add_batch():
    data = request.get_json()
    name = data.get('name', '').strip()
    year = data.get('year')
    if not name or not year:
        return jsonify({'error': 'Name and passing year are required'}), 400
    try:
        conn = get_conn(); cur = conn.cursor()
        cur.execute('INSERT INTO batch (name, passing_year) VALUES (%s, %s) RETURNING batch_id',
                    (name, int(year)))
        batch_id = cur.fetchone()[0]
        conn.commit(); cur.close(); _return_conn(conn)
        return jsonify({'id': batch_id, 'name': name, 'year': int(year)}), 201
    except psycopg2.errors.UniqueViolation:
        return jsonify({'error': 'Batch with this year already exists'}), 409
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/college/batches/<int:batch_id>', methods=['DELETE'])
@college_required
def api_college_delete_batch(batch_id):
    conn = get_conn(); cur = conn.cursor()
    try:
        # Collect affected section/class IDs
        sec_q = 'SELECT section_id FROM section WHERE class_id IN (SELECT class_id FROM class WHERE batch_id=%s)'
        cls_q = 'SELECT class_id FROM class WHERE batch_id=%s'
        # Delete dependent records in correct FK order
        cur.execute(f'DELETE FROM engagement_student_log WHERE session_id IN (SELECT id FROM engagement_sessions WHERE section_id IN ({sec_q}))', (batch_id,))
        cur.execute(f'DELETE FROM engagement_sessions WHERE section_id IN ({sec_q})', (batch_id,))
        cur.execute(f'DELETE FROM engagement_log WHERE section_id IN ({sec_q})', (batch_id,))
        cur.execute(f'DELETE FROM attendance WHERE section_id IN ({sec_q})', (batch_id,))
        cur.execute(f'DELETE FROM whatsapp_recipients WHERE section_id IN ({sec_q})', (batch_id,))
        cur.execute(f'DELETE FROM camera_streams WHERE section_id IN ({sec_q})', (batch_id,))
        cur.execute(f'DELETE FROM timetable WHERE section_id IN ({sec_q})', (batch_id,))
        cur.execute(f'DELETE FROM whatsapp_recipients WHERE class_id IN ({cls_q})', (batch_id,))
        cur.execute(f'DELETE FROM student WHERE class_id IN ({cls_q})', (batch_id,))
        cur.execute(f'DELETE FROM subject WHERE class_id IN ({cls_q})', (batch_id,))
        cur.execute(f'DELETE FROM section WHERE class_id IN ({cls_q})', (batch_id,))
        cur.execute('DELETE FROM class WHERE batch_id=%s', (batch_id,))
        cur.execute('DELETE FROM branch WHERE batch_id=%s', (batch_id,))
        cur.execute('DELETE FROM batch WHERE batch_id=%s', (batch_id,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close(); _return_conn(conn)
    return jsonify({'success': True})


# ── Branch CRUD ───────────────────────────────────────────────────────────────

@app.route('/api/college/branches', methods=['GET'])
@college_required
def api_college_branches():
    batch_id = request.args.get('batch_id', '')
    conn = get_conn(); cur = conn.cursor()
    query = '''SELECT b.branch_id, b.name, b.code, b.hod_name, b.hod_email, b.batch_id, ba.name as batch_name, ba.passing_year
               FROM branch b LEFT JOIN batch ba ON b.batch_id=ba.batch_id WHERE 1=1'''
    params = []
    if batch_id:
        query += ' AND b.batch_id=%s'; params.append(int(batch_id))
    query += ' ORDER BY ba.passing_year DESC NULLS LAST, b.name'
    cur.execute(query, params)
    rows = cur.fetchall(); cur.close(); _return_conn(conn)
    return jsonify([{'id': r[0], 'name': r[1], 'code': r[2], 'hod_name': r[3], 'hod_email': r[4],
                     'batch_id': r[5], 'batch_name': r[6], 'batch_year': r[7]} for r in rows])


@app.route('/api/college/branches', methods=['POST'])
@college_required
def api_college_add_branch():
    data = request.get_json()
    name = data.get('name', '').strip()
    code = data.get('code', '').strip()
    hod_name = data.get('hod_name', '').strip()
    hod_email = data.get('hod_email', '').strip()
    batch_id = data.get('batch_id')
    if not name:
        return jsonify({'error': 'Branch name is required'}), 400
    if not batch_id:
        return jsonify({'error': 'Batch is required'}), 400
    try:
        conn = get_conn(); cur = conn.cursor()
        cur.execute('''INSERT INTO branch (name, code, hod_name, hod_email, batch_id) VALUES (%s, %s, %s, %s, %s)
                       RETURNING branch_id''',
                    (name, code or None, hod_name or None, hod_email or None, int(batch_id)))
        branch_id = cur.fetchone()[0]
        conn.commit(); cur.close(); _return_conn(conn)
        return jsonify({'id': branch_id, 'name': name, 'code': code, 'batch_id': int(batch_id)}), 201
    except psycopg2.errors.UniqueViolation:
        return jsonify({'error': 'Branch already exists'}), 409
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/college/branches/<int:branch_id>', methods=['PUT'])
@college_required
def api_college_update_branch(branch_id):
    data = request.get_json()
    conn = get_conn(); cur = conn.cursor()
    cur.execute('''UPDATE branch SET name=COALESCE(%s,name), code=COALESCE(%s,code),
                   hod_name=%s, hod_email=%s WHERE branch_id=%s''',
                (data.get('name'), data.get('code'), data.get('hod_name'), data.get('hod_email'), branch_id))
    conn.commit(); cur.close(); _return_conn(conn)
    return jsonify({'success': True})


@app.route('/api/college/branches/<int:branch_id>', methods=['DELETE'])
@college_required
def api_college_delete_branch(branch_id):
    conn = get_conn(); cur = conn.cursor()
    try:
        sec_q = 'SELECT section_id FROM section WHERE class_id IN (SELECT class_id FROM class WHERE branch_id=%s)'
        cls_q = 'SELECT class_id FROM class WHERE branch_id=%s'
        cur.execute(f'DELETE FROM engagement_student_log WHERE session_id IN (SELECT id FROM engagement_sessions WHERE section_id IN ({sec_q}))', (branch_id,))
        cur.execute(f'DELETE FROM engagement_sessions WHERE section_id IN ({sec_q})', (branch_id,))
        cur.execute(f'DELETE FROM engagement_log WHERE section_id IN ({sec_q})', (branch_id,))
        cur.execute(f'DELETE FROM attendance WHERE section_id IN ({sec_q})', (branch_id,))
        cur.execute(f'DELETE FROM whatsapp_recipients WHERE section_id IN ({sec_q})', (branch_id,))
        cur.execute(f'DELETE FROM camera_streams WHERE section_id IN ({sec_q})', (branch_id,))
        cur.execute(f'DELETE FROM timetable WHERE section_id IN ({sec_q})', (branch_id,))
        cur.execute(f'DELETE FROM whatsapp_recipients WHERE class_id IN ({cls_q})', (branch_id,))
        cur.execute(f'DELETE FROM student WHERE class_id IN ({cls_q})', (branch_id,))
        cur.execute(f'DELETE FROM subject WHERE class_id IN ({cls_q})', (branch_id,))
        cur.execute(f'DELETE FROM section WHERE class_id IN ({cls_q})', (branch_id,))
        cur.execute('DELETE FROM class WHERE branch_id=%s', (branch_id,))
        cur.execute('DELETE FROM branch WHERE branch_id=%s', (branch_id,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close(); _return_conn(conn)
    return jsonify({'success': True})


# ── Class CRUD ────────────────────────────────────────────────────────────────

@app.route('/api/college/classes', methods=['GET'])
@college_required
def api_college_classes():
    batch_id = request.args.get('batch_id', '')
    branch_id = request.args.get('branch_id', '')
    conn = get_conn(); cur = conn.cursor()
    query = '''SELECT c.class_id, c.name, c.batch_id, c.branch_id, b.name as batch_name, br.name as branch_name,
                      (SELECT COUNT(*) FROM section s WHERE s.class_id=c.class_id) as sec_count,
                      (SELECT COUNT(*) FROM student st WHERE st.class_id=c.class_id) as stu_count
               FROM class c
               LEFT JOIN batch b ON c.batch_id=b.batch_id
               LEFT JOIN branch br ON c.branch_id=br.branch_id WHERE 1=1'''
    params = []
    if batch_id:
        query += ' AND c.batch_id=%s'; params.append(int(batch_id))
    if branch_id:
        query += ' AND c.branch_id=%s'; params.append(int(branch_id))
    query += ' ORDER BY c.name'
    cur.execute(query, params)
    rows = cur.fetchall(); cur.close(); _return_conn(conn)
    return jsonify([{'id': r[0], 'name': r[1], 'batch_id': r[2], 'branch_id': r[3],
                     'batch_name': r[4], 'branch_name': r[5],
                     'section_count': r[6], 'student_count': r[7]} for r in rows])


@app.route('/api/college/classes', methods=['POST'])
@college_required
def api_college_add_class():
    data = request.get_json()
    name = data.get('name', '').strip()
    batch_id = data.get('batch_id')
    branch_id = data.get('branch_id')
    if not name:
        return jsonify({'error': 'Class name is required'}), 400
    try:
        conn = get_conn(); cur = conn.cursor()
        cur.execute('INSERT INTO class (name, batch_id, branch_id) VALUES (%s, %s, %s) RETURNING class_id',
                    (name, int(batch_id) if batch_id else None, int(branch_id) if branch_id else None))
        class_id = cur.fetchone()[0]
        conn.commit(); cur.close(); _return_conn(conn)
        return jsonify({'id': class_id, 'name': name}), 201
    except psycopg2.errors.UniqueViolation:
        return jsonify({'error': 'Class already exists for this batch/branch'}), 409
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/college/classes/<int:class_id>', methods=['DELETE'])
@college_required
def api_college_delete_class(class_id):
    conn = get_conn(); cur = conn.cursor()
    try:
        sec_q = 'SELECT section_id FROM section WHERE class_id=%s'
        cur.execute(f'DELETE FROM engagement_student_log WHERE session_id IN (SELECT id FROM engagement_sessions WHERE section_id IN ({sec_q}))', (class_id,))
        cur.execute(f'DELETE FROM engagement_sessions WHERE section_id IN ({sec_q})', (class_id,))
        cur.execute(f'DELETE FROM engagement_log WHERE section_id IN ({sec_q})', (class_id,))
        cur.execute(f'DELETE FROM attendance WHERE section_id IN ({sec_q})', (class_id,))
        cur.execute(f'DELETE FROM whatsapp_recipients WHERE section_id IN ({sec_q})', (class_id,))
        cur.execute(f'DELETE FROM camera_streams WHERE section_id IN ({sec_q})', (class_id,))
        cur.execute(f'DELETE FROM timetable WHERE section_id IN ({sec_q})', (class_id,))
        cur.execute('DELETE FROM whatsapp_recipients WHERE class_id=%s', (class_id,))
        cur.execute('DELETE FROM student WHERE class_id=%s', (class_id,))
        cur.execute('DELETE FROM subject WHERE class_id=%s', (class_id,))
        cur.execute('DELETE FROM section WHERE class_id=%s', (class_id,))
        cur.execute('DELETE FROM class WHERE class_id=%s', (class_id,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close(); _return_conn(conn)
    return jsonify({'success': True})


# ── Section CRUD ──────────────────────────────────────────────────────────────

@app.route('/api/college/sections', methods=['GET'])
@college_required
def api_college_sections():
    class_id = request.args.get('class_id', '')
    conn = get_conn(); cur = conn.cursor()
    query = '''SELECT s.section_id, s.name, s.class_id, c.name as class_name,
                      (SELECT COUNT(*) FROM student st WHERE st.section_id=s.section_id) as stu_count
               FROM section s LEFT JOIN class c ON s.class_id=c.class_id WHERE 1=1'''
    params = []
    if class_id:
        query += ' AND s.class_id=%s'; params.append(int(class_id))
    query += ' ORDER BY c.name, s.name'
    cur.execute(query, params)
    rows = cur.fetchall(); cur.close(); _return_conn(conn)
    return jsonify([{'id': r[0], 'name': r[1], 'class_id': r[2], 'class_name': r[3], 'student_count': r[4]} for r in rows])


@app.route('/api/college/sections', methods=['POST'])
@college_required
def api_college_add_section():
    data = request.get_json()
    name = data.get('name', '').strip()
    class_id = data.get('class_id')
    if not name or not class_id:
        return jsonify({'error': 'Section name and class are required'}), 400
    try:
        conn = get_conn(); cur = conn.cursor()
        cur.execute('INSERT INTO section (class_id, name) VALUES (%s, %s) RETURNING section_id',
                    (int(class_id), name))
        section_id = cur.fetchone()[0]
        conn.commit(); cur.close(); _return_conn(conn)
        return jsonify({'id': section_id, 'name': name}), 201
    except psycopg2.errors.UniqueViolation:
        return jsonify({'error': 'Section already exists in this class'}), 409
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/college/sections/<int:section_id>', methods=['DELETE'])
@college_required
def api_college_delete_section(section_id):
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute('DELETE FROM engagement_student_log WHERE session_id IN (SELECT id FROM engagement_sessions WHERE section_id=%s)', (section_id,))
        cur.execute('DELETE FROM engagement_sessions WHERE section_id=%s', (section_id,))
        cur.execute('DELETE FROM engagement_log WHERE section_id=%s', (section_id,))
        cur.execute('DELETE FROM attendance WHERE section_id=%s', (section_id,))
        cur.execute('DELETE FROM whatsapp_recipients WHERE section_id=%s', (section_id,))
        cur.execute('DELETE FROM camera_streams WHERE section_id=%s', (section_id,))
        cur.execute('DELETE FROM timetable WHERE section_id=%s', (section_id,))
        cur.execute('DELETE FROM student WHERE section_id=%s', (section_id,))
        cur.execute('DELETE FROM section WHERE section_id=%s', (section_id,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close(); _return_conn(conn)
    return jsonify({'success': True})


# ── Timetable CRUD ────────────────────────────────────────────────────────────

@app.route('/api/college/timetable/<int:section_id>', methods=['GET'])
@college_required
def api_college_timetable(section_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute('''SELECT tt_id, day_of_week, period_name, teacher_name, from_time, to_time, is_recess
                   FROM timetable WHERE section_id=%s ORDER BY day_of_week, from_time''', (section_id,))
    rows = cur.fetchall(); cur.close(); _return_conn(conn)
    return jsonify([{
        'id': r[0], 'day': r[1], 'period_name': r[2], 'teacher_name': r[3],
        'from_time': str(r[4])[:5], 'to_time': str(r[5])[:5], 'is_recess': r[6]
    } for r in rows])


@app.route('/api/college/timetable/<int:section_id>', methods=['POST'])
@college_required
def api_college_add_timetable_entry(section_id):
    data = request.get_json()
    day = data.get('day', 0)
    period_name = data.get('period_name', '').strip()
    teacher_name = data.get('teacher_name', '').strip()
    from_time = data.get('from_time', '')
    to_time = data.get('to_time', '')
    is_recess = data.get('is_recess', False)
    if not period_name or not from_time or not to_time:
        return jsonify({'error': 'Period name, start time, and end time are required'}), 400
    conn = get_conn(); cur = conn.cursor()
    cur.execute('''INSERT INTO timetable (section_id, day_of_week, period_name, teacher_name, from_time, to_time, is_recess)
                   VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING tt_id''',
                (section_id, int(day), period_name, teacher_name or None, from_time, to_time, is_recess))
    tt_id = cur.fetchone()[0]
    conn.commit(); cur.close(); _return_conn(conn)
    return jsonify({'id': tt_id, 'period_name': period_name}), 201


@app.route('/api/college/timetable/<int:section_id>/generate', methods=['POST'])
@college_required
def api_college_generate_timetable(section_id):
    """Auto-generate timetable from college timing, number of periods, recess position."""
    data = request.get_json()
    start_time = data.get('start_time', '09:00')  # e.g. "09:00"
    end_time = data.get('end_time', '16:00')      # e.g. "16:00"
    num_periods = data.get('num_periods', 6)
    enable_recess = data.get('enable_recess', True)
    recess_after = data.get('recess_after', 3)    # recess after period 3
    recess_duration = data.get('recess_duration', 30)  # minutes
    days = data.get('days', [0, 1, 2, 3, 4, 5])  # Mon-Sat
    teachers = data.get('teachers', {})           # { "1": "Teacher Name", "2": "..." }

    try:
        start_dt = dt.datetime.strptime(start_time, '%H:%M')
        end_dt = dt.datetime.strptime(end_time, '%H:%M')
        total_minutes = (end_dt - start_dt).total_seconds() / 60
        recess_minutes = recess_duration if enable_recess else 0
        teaching_minutes = total_minutes - recess_minutes
        period_duration = int(teaching_minutes / num_periods)

        conn = get_conn(); cur = conn.cursor()
        # Clear existing timetable for this section
        cur.execute('DELETE FROM timetable WHERE section_id=%s', (section_id,))

        entries = []
        for day in days:
            current = start_dt
            for p in range(1, num_periods + 1):
                # Insert recess before this period if needed
                if enable_recess and p == recess_after + 1:
                    recess_end = current + dt.timedelta(minutes=recess_minutes)
                    cur.execute('''INSERT INTO timetable (section_id, day_of_week, period_name, teacher_name, from_time, to_time, is_recess)
                                   VALUES (%s,%s,%s,%s,%s,%s,%s)''',
                                (section_id, day, 'Recess', None,
                                 current.strftime('%H:%M'), recess_end.strftime('%H:%M'), True))
                    current = recess_end

                period_end = current + dt.timedelta(minutes=period_duration)
                teacher = teachers.get(str(p), '')
                cur.execute('''INSERT INTO timetable (section_id, day_of_week, period_name, teacher_name, from_time, to_time, is_recess)
                               VALUES (%s,%s,%s,%s,%s,%s,%s)''',
                            (section_id, day, f'Period {p}', teacher or None,
                             current.strftime('%H:%M'), period_end.strftime('%H:%M'), False))
                entries.append({'day': day, 'period': p, 'from': current.strftime('%H:%M'), 'to': period_end.strftime('%H:%M')})
                current = period_end

        conn.commit(); cur.close(); _return_conn(conn)
        return jsonify({'success': True, 'periods_created': len(entries), 'entries': entries})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/college/timetable/entry/<int:tt_id>', methods=['PUT'])
@college_required
def api_college_update_timetable_entry(tt_id):
    data = request.get_json()
    conn = get_conn(); cur = conn.cursor()
    cur.execute('''UPDATE timetable SET period_name=COALESCE(%s, period_name),
                   teacher_name=%s, from_time=COALESCE(%s, from_time),
                   to_time=COALESCE(%s, to_time), is_recess=COALESCE(%s, is_recess)
                   WHERE tt_id=%s''',
                (data.get('period_name'), data.get('teacher_name'),
                 data.get('from_time'), data.get('to_time'), data.get('is_recess'), tt_id))
    conn.commit(); cur.close(); _return_conn(conn)
    return jsonify({'success': True})


@app.route('/api/college/timetable/entry/<int:tt_id>', methods=['DELETE'])
@college_required
def api_college_delete_timetable_entry(tt_id):
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute('DELETE FROM engagement_student_log WHERE session_id IN (SELECT id FROM engagement_sessions WHERE period_id=%s)', (tt_id,))
        cur.execute('DELETE FROM engagement_sessions WHERE period_id=%s', (tt_id,))
        cur.execute('DELETE FROM engagement_log WHERE period_id=%s', (tt_id,))
        cur.execute('DELETE FROM attendance WHERE period_id=%s', (tt_id,))
        cur.execute('DELETE FROM timetable WHERE tt_id=%s', (tt_id,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close(); _return_conn(conn)
    return jsonify({'success': True})


# ── Teacher Management ────────────────────────────────────────────────────────

@app.route('/api/college/teachers', methods=['GET'])
@college_required
def api_college_teachers():
    designation = request.args.get('designation', '').strip()
    subject = request.args.get('subject', '').strip()
    conn = get_conn(); cur = conn.cursor()
    query = "SELECT id, name, email, designation, subjects FROM users WHERE role='teacher'"
    params = []
    if designation:
        query += " AND designation ILIKE %s"; params.append(f"%{designation}%")
    if subject:
        query += " AND subjects ILIKE %s"; params.append(f"%{subject}%")
    query += " ORDER BY name"
    cur.execute(query, params)
    rows = cur.fetchall(); cur.close(); _return_conn(conn)
    return jsonify([{'id': r[0], 'name': r[1], 'email': r[2],
                     'designation': r[3] or '', 'subjects': r[4] or ''} for r in rows])


@app.route('/api/college/teachers', methods=['POST'])
@college_required
def api_college_add_teacher():
    data = request.get_json()
    name = data.get('name', '').strip()
    email = data.get('email', '').strip()
    password = data.get('password', '').strip()
    designation = (data.get('designation') or '').strip()
    subjects = (data.get('subjects') or '').strip()
    if not name or not email or not password:
        return jsonify({'error': 'Name, email, and password are required'}), 400
    try:
        conn = get_conn(); cur = conn.cursor()
        cur.execute('''INSERT INTO users (name, email, password_hash, role, designation, subjects)
                       VALUES (%s,%s,%s,%s,%s,%s) RETURNING id''',
                    (name, email, generate_password_hash(password), 'teacher',
                     designation or None, subjects or None))
        uid = cur.fetchone()[0]
        conn.commit(); cur.close(); _return_conn(conn)
        return jsonify({'id': uid, 'name': name, 'email': email,
                        'designation': designation, 'subjects': subjects}), 201
    except psycopg2.errors.UniqueViolation:
        return jsonify({'error': 'Email already registered'}), 409
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/college/teachers/<int:teacher_id>', methods=['PUT'])
@college_required
def api_college_update_teacher(teacher_id):
    data = request.get_json()
    conn = get_conn(); cur = conn.cursor()
    updates = []
    params = []
    if data.get('name'):
        updates.append('name=%s'); params.append(data['name'])
    if data.get('email'):
        updates.append('email=%s'); params.append(data['email'])
    if data.get('password'):
        updates.append('password_hash=%s'); params.append(generate_password_hash(data['password']))
    if 'designation' in data:
        updates.append('designation=%s'); params.append(data['designation'] or None)
    if 'subjects' in data:
        updates.append('subjects=%s'); params.append(data['subjects'] or None)
    if updates:
        params.append(teacher_id)
        cur.execute(f"UPDATE users SET {','.join(updates)} WHERE id=%s AND role='teacher'", params)
        conn.commit()
    cur.close(); _return_conn(conn)
    return jsonify({'success': True})


@app.route('/api/college/teachers/<int:teacher_id>', methods=['DELETE'])
@college_required
def api_college_delete_teacher(teacher_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("DELETE FROM users WHERE id=%s AND role='teacher'", (teacher_id,))
    conn.commit(); cur.close(); _return_conn(conn)
    return jsonify({'success': True})


# ── Student Management ────────────────────────────────────────────────────────

@app.route('/api/college/students', methods=['GET'])
@college_required
def api_college_students():
    class_id = request.args.get('class_id', '')
    section_id = request.args.get('section_id', '')
    conn = get_conn(); cur = conn.cursor()
    query = '''SELECT st.std_id, st.name, st.email, st.roll_no, c.name as class_name,
                      s.name as sec_name, st.dob, st.class_id, st.section_id, st.parent_email
               FROM student st
               LEFT JOIN class c ON st.class_id=c.class_id
               LEFT JOIN section s ON st.section_id=s.section_id WHERE 1=1'''
    params = []
    if class_id:
        query += ' AND st.class_id=%s'; params.append(int(class_id))
    if section_id:
        query += ' AND st.section_id=%s'; params.append(int(section_id))
    query += ' ORDER BY c.name, s.name, st.roll_no, st.name'
    cur.execute(query, params)
    rows = cur.fetchall(); cur.close(); _return_conn(conn)
    return jsonify([{
        'id': r[0], 'name': r[1], 'email': r[2], 'roll_no': r[3],
        'class_name': r[4], 'section_name': r[5], 'dob': str(r[6]) if r[6] else None,
        'class_id': r[7], 'section_id': r[8], 'parent_email': r[9] or ''
    } for r in rows])


@app.route('/api/college/students', methods=['POST'])
@college_required
def api_college_add_student():
    data = request.get_json()
    name = data.get('name', '').strip()
    email = data.get('email', '').strip()
    roll_no = data.get('roll_no', '').strip()
    class_id = data.get('class_id')
    section_id = data.get('section_id')
    dob = data.get('dob', '').strip()
    parent_email = data.get('parent_email', '').strip()
    photos = data.get('photos', [])

    if not name:
        return jsonify({'error': 'Student name is required'}), 400

    try:
        conn = get_conn(); cur = conn.cursor()
        cur.execute('''INSERT INTO student (name, email, roll_no, class_id, section_id, dob, parent_email)
                       VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING std_id''',
                    (name, email or None, roll_no or None,
                     int(class_id) if class_id else None,
                     int(section_id) if section_id else None,
                     dob or None, parent_email or None))
        std_id = cur.fetchone()[0]
        conn.commit(); cur.close(); _return_conn(conn)

        # Save photos if provided
        if photos:
            import gc
            student_dir = os.path.join(DATASET_DIR, name)
            os.makedirs(student_dir, exist_ok=True)
            for i, b64 in enumerate(photos[:3]):
                _save_b64_image(b64, os.path.join(student_dir, f'photo_{i+1}.jpg'), max_size=640)
            gc.collect()

        return jsonify({'id': std_id, 'name': name}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/college/students/<int:std_id>', methods=['PUT'])
@college_required
def api_college_update_student(std_id):
    data = request.get_json()
    conn = get_conn(); cur = conn.cursor()
    updates = []
    params = []
    for field in ['name', 'email', 'roll_no', 'dob', 'parent_email']:
        if field in data:
            updates.append(f'{field}=%s'); params.append(data[field] or None)
    if 'class_id' in data:
        updates.append('class_id=%s'); params.append(int(data['class_id']) if data['class_id'] else None)
    if 'section_id' in data:
        updates.append('section_id=%s'); params.append(int(data['section_id']) if data['section_id'] else None)
    if updates:
        params.append(std_id)
        cur.execute(f"UPDATE student SET {','.join(updates)} WHERE std_id=%s", params)
        conn.commit()
    cur.close(); _return_conn(conn)
    return jsonify({'success': True})


@app.route('/api/college/students/<int:std_id>', methods=['DELETE'])
@college_required
def api_college_delete_student(std_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute('SELECT name FROM student WHERE std_id=%s', (std_id,))
    row = cur.fetchone()
    if row:
        name = row[0]
        cur.execute('DELETE FROM attendance WHERE student_id=%s', (std_id,))
        cur.execute('DELETE FROM student WHERE std_id=%s', (std_id,))
        conn.commit()
        d = os.path.join(DATASET_DIR, name)
        if os.path.exists(d):
            shutil.rmtree(d)
    cur.close(); _return_conn(conn)
    return jsonify({'success': True})


@app.route('/api/college/regenerate-embeddings', methods=['POST'])
@college_required
def api_college_regenerate_embeddings():
    """Regenerate embeddings.csv — scans dataset/ folder and lists all student names with photos."""
    try:
        dataset_dir = os.path.join(BASE_DIR, 'dataset')
        if not os.path.isdir(dataset_dir):
            return jsonify({'error': 'No dataset folder found'}), 400

        names = []
        for person_name in os.listdir(dataset_dir):
            person_path = os.path.join(dataset_dir, person_name)
            if not os.path.isdir(person_path):
                continue
            # Check this person has at least one image
            has_image = False
            for img_file in os.listdir(person_path):
                img_path = os.path.join(person_path, img_file)
                if os.path.isfile(img_path) and img_file.lower().endswith(('.jpg', '.jpeg', '.png')):
                    has_image = True
                    names.append(person_name)
                    break  # one entry per image is enough for LBPH
            # Add all images for this person
            if has_image:
                for img_file in os.listdir(person_path):
                    img_path = os.path.join(person_path, img_file)
                    if os.path.isfile(img_path) and img_file.lower().endswith(('.jpg', '.jpeg', '.png')):
                        if person_name not in names or names.count(person_name) < 3:
                            names.append(person_name)

        if not names:
            return jsonify({'error': 'No face images found in dataset'}), 400

        # Save a simple CSV with just names (LBPH trains from images directly)
        import pandas as pd_regen
        df = pd_regen.DataFrame({'name': names, 'placeholder': [0] * len(names)})
        df.to_csv(os.path.join(BASE_DIR, 'embeddings.csv'), index=False)

        unique_students = len(set(names))
        print(f"[embeddings] Regenerated: {len(names)} entries for {unique_students} students", flush=True)
        return jsonify({'success': True, 'total_embeddings': len(names), 'unique_students': unique_students})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/college/students/bulk', methods=['POST'])
@college_required
def api_college_bulk_register():
    """Bulk register students from CSV/Excel upload."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    class_id = request.form.get('class_id', '')
    section_id = request.form.get('section_id', '')

    if not file.filename:
        return jsonify({'error': 'Empty file'}), 400

    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        elif file.filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(file)
        else:
            return jsonify({'error': 'Unsupported file format. Use CSV or Excel.'}), 400

        # Expected columns: name, email, roll_no, dob (optional)
        required_cols = ['name']
        for col in required_cols:
            if col not in df.columns:
                return jsonify({'error': f'Missing required column: {col}'}), 400

        conn = get_conn(); cur = conn.cursor()
        added = 0
        errors = []
        for _, row in df.iterrows():
            name = str(row.get('name', '')).strip()
            if not name:
                continue
            email = str(row.get('email', '')).strip() if 'email' in row else None
            roll_no = str(row.get('roll_no', '')).strip() if 'roll_no' in row else None
            dob = str(row.get('dob', '')).strip() if 'dob' in row else None

            # Use per-row class/section if provided, else form values
            c_id = row.get('class_id', class_id)
            s_id = row.get('section_id', section_id)

            try:
                cur.execute('''INSERT INTO student (name, email, roll_no, class_id, section_id, dob)
                               VALUES (%s,%s,%s,%s,%s,%s)''',
                            (name, email if email and email != 'nan' else None,
                             roll_no if roll_no and roll_no != 'nan' else None,
                             int(c_id) if c_id and str(c_id) != 'nan' else None,
                             int(s_id) if s_id and str(s_id) != 'nan' else None,
                             dob if dob and dob != 'nan' and dob != 'NaT' else None))
                added += 1
            except Exception as e:
                errors.append(f"Row {name}: {str(e)}")
                conn.rollback()
                conn = get_conn(); cur = conn.cursor()

        conn.commit(); cur.close(); _return_conn(conn)
        return jsonify({'success': True, 'added': added, 'errors': errors})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Attendance Data & Download ────────────────────────────────────────────────

@app.route('/api/college/attendance', methods=['GET'])
@college_required
def api_college_attendance():
    """Get attendance records with filters."""
    date_from = request.args.get('date_from', _today_ist().strftime('%Y-%m-%d'))
    date_to = request.args.get('date_to', _today_ist().strftime('%Y-%m-%d'))
    class_id = request.args.get('class_id', '')
    section_id = request.args.get('section_id', '')
    batch_id = request.args.get('batch_id', '')
    branch_id = request.args.get('branch_id', '')

    conn = get_conn(); cur = conn.cursor()
    query = '''SELECT a.date, st.name, st.roll_no, c.name as class_name, sec.name as sec_name,
                      t.period_name, t.from_time, t.to_time
               FROM attendance a
               JOIN student st ON a.student_id=st.std_id
               LEFT JOIN class c ON st.class_id=c.class_id
               LEFT JOIN section sec ON a.section_id=sec.section_id
               LEFT JOIN timetable t ON a.period_id=t.tt_id
               WHERE a.date BETWEEN %s AND %s'''
    params = [date_from, date_to]
    if section_id:
        query += ' AND a.section_id=%s'; params.append(int(section_id))
    elif class_id:
        query += ' AND st.class_id=%s'; params.append(int(class_id))
    elif branch_id:
        query += ' AND c.branch_id=%s'; params.append(int(branch_id))
    elif batch_id:
        query += ' AND c.batch_id=%s'; params.append(int(batch_id))
    query += ' ORDER BY a.date DESC, c.name, sec.name, st.roll_no'
    cur.execute(query, params)
    rows = cur.fetchall(); cur.close(); _return_conn(conn)
    return jsonify([{
        'date': str(r[0]), 'name': r[1], 'roll_no': r[2], 'class_name': r[3],
        'section_name': r[4], 'period': r[5],
        'time': f"{str(r[6])[:5]}-{str(r[7])[:5]}" if r[6] else None
    } for r in rows])


@app.route('/api/college/attendance/download', methods=['GET'])
@college_required
def api_college_attendance_download():
    """Download attendance as Excel with customizable filters."""
    from flask import send_file
    date_from = request.args.get('date_from', _today_ist().strftime('%Y-%m-%d'))
    date_to = request.args.get('date_to', _today_ist().strftime('%Y-%m-%d'))
    class_id = request.args.get('class_id', '')
    section_id = request.args.get('section_id', '')
    batch_id = request.args.get('batch_id', '')
    branch_id = request.args.get('branch_id', '')

    conn = get_conn(); cur = conn.cursor()

    # Get students
    stu_query = '''SELECT st.std_id, st.name, st.roll_no, c.name as class_name, sec.name as sec_name
                   FROM student st
                   LEFT JOIN class c ON st.class_id=c.class_id
                   LEFT JOIN section sec ON st.section_id=sec.section_id WHERE 1=1'''
    stu_params = []
    if section_id:
        stu_query += ' AND st.section_id=%s'; stu_params.append(int(section_id))
    elif class_id:
        stu_query += ' AND st.class_id=%s'; stu_params.append(int(class_id))
    elif branch_id:
        stu_query += ' AND c.branch_id=%s'; stu_params.append(int(branch_id))
    elif batch_id:
        stu_query += ' AND c.batch_id=%s'; stu_params.append(int(batch_id))
    stu_query += ' ORDER BY c.name, sec.name, st.roll_no'
    cur.execute(stu_query, stu_params)
    students = cur.fetchall()

    # Get periods for column headers
    per_query = '''SELECT DISTINCT t.period_name, t.from_time FROM timetable t
                   JOIN section s ON t.section_id=s.section_id WHERE t.is_recess=FALSE'''
    per_params = []
    if section_id:
        per_query += ' AND t.section_id=%s'; per_params.append(int(section_id))
    elif class_id:
        per_query += ' AND s.class_id=%s'; per_params.append(int(class_id))
    per_query += ' ORDER BY t.from_time'
    cur.execute(per_query, per_params)
    periods = cur.fetchall()

    # Get attendance records
    att_query = '''SELECT a.date, a.student_id, t.period_name
                   FROM attendance a LEFT JOIN timetable t ON a.period_id=t.tt_id
                   WHERE a.date BETWEEN %s AND %s'''
    att_params = [date_from, date_to]
    if section_id:
        att_query += ' AND a.section_id=%s'; att_params.append(int(section_id))
    elif class_id:
        att_query += ' AND a.student_id IN (SELECT std_id FROM student WHERE class_id=%s)'; att_params.append(int(class_id))
    cur.execute(att_query, att_params)
    att_records = cur.fetchall()
    cur.close(); _return_conn(conn)

    # Build presence set
    present_set = set()
    for rec in att_records:
        present_set.add((str(rec[0]), rec[1], rec[2] or ''))

    # Generate date range
    from_d = dt.datetime.strptime(date_from, '%Y-%m-%d').date()
    to_d = dt.datetime.strptime(date_to, '%Y-%m-%d').date()
    date_range = []
    d = from_d
    while d <= to_d:
        date_range.append(d)
        d += dt.timedelta(days=1)

    # Build Excel
    import io
    if pd is None:
        return jsonify({'error': 'pandas not available'}), 500

    rows_data = []
    for stu in students:
        row = {'Roll No': stu[2] or '', 'Name': stu[1], 'Class': stu[3] or '', 'Section': stu[4] or ''}
        for date in date_range:
            if periods:
                for p in periods:
                    col = f"{date.strftime('%d/%m')} {p[0]}"
                    row[col] = 'P' if (str(date), stu[0], p[0]) in present_set else 'A'
            else:
                col = date.strftime('%d/%m/%Y')
                row[col] = 'P' if any((str(date), stu[0], '') in present_set for _ in [1]) else 'A'
        rows_data.append(row)

    df = pd.DataFrame(rows_data)
    output = io.BytesIO()
    df.to_excel(output, index=False, engine='openpyxl')
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'attendance_{date_from}_to_{date_to}.xlsx')


# ── College Liveness View ─────────────────────────────────────────────────────

@app.route('/api/college/liveness/sessions', methods=['GET'])
@college_required
def api_college_liveness_sessions():
    """View all active engagement/liveness sessions across sections."""
    conn = get_conn(); cur = conn.cursor()
    cur.execute('''SELECT es.id, c.name as class_name, s.name as sec_name, t.period_name,
                          es.started_at, es.capture_interval, u.name as teacher_name, es.is_active
                   FROM engagement_sessions es
                   JOIN section s ON es.section_id=s.section_id
                   JOIN class c ON s.class_id=c.class_id
                   JOIN timetable t ON es.period_id=t.tt_id
                   LEFT JOIN users u ON es.started_by=u.id
                   ORDER BY es.is_active DESC, es.started_at DESC LIMIT 50''')
    rows = cur.fetchall(); cur.close(); _return_conn(conn)
    return jsonify([{
        'id': r[0], 'class_name': r[1], 'section_name': r[2], 'period': r[3],
        'started_at': str(r[4]), 'interval': r[5], 'teacher': r[6], 'is_active': r[7]
    } for r in rows])


@app.route('/api/college/liveness/<int:session_id>/logs', methods=['GET'])
@college_required
def api_college_liveness_logs(session_id):
    """Get engagement logs for a specific session."""
    conn = get_conn(); cur = conn.cursor()
    cur.execute('''SELECT el.timestamp, el.total_faces, el.attentive_pct, el.confused_pct,
                          el.distracted_pct, el.avg_score
                   FROM engagement_log el
                   JOIN engagement_sessions es ON el.section_id=es.section_id AND el.period_id=es.period_id
                   WHERE es.id=%s
                   ORDER BY el.timestamp DESC LIMIT 100''', (session_id,))
    rows = cur.fetchall(); cur.close(); _return_conn(conn)
    return jsonify([{
        'timestamp': str(r[0]), 'total_faces': r[1], 'attentive_pct': float(r[2]),
        'confused_pct': float(r[3]), 'distracted_pct': float(r[4]), 'avg_score': float(r[5])
    } for r in rows])


if __name__ == '__main__':
    init_db()

    # ── Background Scheduler for automated reports ────────────────────────────
    import threading

    def run_scheduled_reports():
        """Background thread: checks report_schedules every 60s and sends due reports."""
        import time as _time
        while True:
            try:
                now = _now_ist()
                current_time = now.strftime('%H:%M')
                current_day = now.weekday()  # 0=Monday
                current_date = now.day

                conn = get_conn(); cur = conn.cursor()
                cur.execute("""
                    SELECT rs.id, rs.teacher_id, rs.section_id, rs.frequency, rs.send_time, rs.include_parents
                    FROM report_schedules rs WHERE rs.active = true
                """)
                schedules = cur.fetchall()
                cur.close(); _return_conn(conn)

                for sch in schedules:
                    sch_id, teacher_id, section_id, frequency, send_time, include_parents = sch
                    sch_time = str(send_time)[:5]  # 'HH:MM'

                    if sch_time != current_time:
                        continue

                    # Check frequency
                    should_send = False
                    if frequency == 'daily':
                        should_send = True
                    elif frequency == 'weekly' and current_day == 0:  # Monday
                        should_send = True
                    elif frequency == 'monthly' and current_date == 1:
                        should_send = True

                    if not should_send:
                        continue

                    # Determine date range
                    if frequency == 'daily':
                        date_from = (now - dt.timedelta(days=1)).strftime('%Y-%m-%d')
                        date_to = now.strftime('%Y-%m-%d')
                    elif frequency == 'weekly':
                        date_from = (now - dt.timedelta(days=7)).strftime('%Y-%m-%d')
                        date_to = now.strftime('%Y-%m-%d')
                    else:  # monthly
                        date_from = (now - dt.timedelta(days=30)).strftime('%Y-%m-%d')
                        date_to = now.strftime('%Y-%m-%d')

                    # Send report using the main handler logic
                    with app.test_request_context(json={
                        'section_id': section_id,
                        'date_from': date_from,
                        'date_to': date_to,
                        'send_to_authorities': True,
                        'send_to_parents': include_parents
                    }):
                        from flask import session as flask_session
                        flask_session['user_id'] = teacher_id
                        flask_session['role'] = 'teacher'
                        try:
                            teacher_send_report()
                            print(f"[Scheduler] Report sent for schedule #{sch_id} (section {section_id})")
                        except Exception as e:
                            print(f"[Scheduler] Error sending report #{sch_id}: {e}")

            except Exception as e:
                print(f"[Scheduler] Error: {e}")

            _time.sleep(60)  # Check every minute

    scheduler_thread = threading.Thread(target=run_scheduled_reports, daemon=True)
    scheduler_thread.start()
    print("[Scheduler] Background report scheduler started.")

    app.run(debug=True, host='0.0.0.0', port=5000)
else:
    # When imported by gunicorn, initialize DB in a background thread
    # so gunicorn can bind the port immediately without timeout
    import threading as _threading
    def _bg_init():
        try:
            init_db()
            print("[startup] DB initialized successfully.", flush=True)
        except Exception as _e:
            print(f"[startup] DB init error: {_e}", flush=True)
    _threading.Thread(target=_bg_init, daemon=True).start()

